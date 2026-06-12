#!/usr/bin/env python3
"""ЗОНД телефонов kufar — диагностика, НЕ массовый сбор (2 объявления).

ЗАПУСКАТЬ С ВЫКЛЮЧЕННЫМ Psiphon (нужен белорусский IP)!
    1) выключите Psiphon;
    2) ./bin/python kufar_phone_probe.py
    3) пришлите весь вывод Claude;
    4) включите Psiphon обратно.

Что делает: открывает деталку kufar в реальном окне браузера, проверяет наш IP/страну,
смотрит, есть ли номер сразу в __NEXT_DATA__, жмёт «Показать телефон», ловит сетевой
запрос к /phone и его ответ, читает раскрытый номер из DOM. По выводу станет ясно,
проходит ли reCAPTCHA автоматически и реально ли строить авто-добор.
"""
import asyncio
import json

from openpyxl import load_workbook
from playwright.async_api import async_playwright

MAIN_XLSX = "commercial_realty.xlsx"
N = 2  # сколько объявлений проверить


def kufar_urls_without_phone(limit):
    wb = load_workbook(MAIN_XLSX, read_only=True)
    out = []
    for sheet in ("Продажа", "Аренда"):
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        rows = ws.iter_rows(values_only=True)
        next(rows, None)
        hdr = next(rows, None)
        if not hdr:
            continue
        col = {str(h): i for i, h in enumerate(hdr) if h}
        ph, lk, sr = col.get("Телефон"), col.get("Ссылка"), col.get("Источник")
        for r in rows:
            if not r or sr is None or "kufar" not in str(r[sr] or ""):
                continue
            if ph is not None and str(r[ph] or "").strip():
                continue
            if lk is not None and str(r[lk] or "").startswith("http"):
                out.append(str(r[lk]))
            if len(out) >= limit:
                wb.close()
                return out
    wb.close()
    return out


async def main():
    urls = kufar_urls_without_phone(N)
    print(f"Проверяю {len(urls)} kufar-объявлений без телефона:\n")

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=False)  # видимое окно — лучше для reCAPTCHA
        ctx = await browser.new_context(
            locale="ru-RU",
            user_agent=("Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                        "AppleWebKit/605.1.15 (KHTML, like Gecko) Version/16.4 Safari/605.1.15"))
        page = await ctx.new_page()

        # 0) наш IP/страна глазами сайта
        try:
            await page.goto("http://ip-api.com/json/?fields=country,query,proxy,hosting",
                            wait_until="load", timeout=15000)
            ipinfo = await page.evaluate("() => document.body.innerText")
            print("НАШ IP сейчас:", ipinfo.strip(), "\n")
        except Exception as e:
            print("IP проверить не вышло:", e, "\n")

        phone_responses = []

        async def on_response(resp):
            if "/phone" in resp.url:
                try:
                    body = await resp.text()
                except Exception:
                    body = "<не прочитал>"
                phone_responses.append((resp.status, resp.url, body[:300]))

        page.on("response", on_response)

        for i, url in enumerate(urls, 1):
            print(f"=== [{i}] {url}")
            phone_responses.clear()
            try:
                await page.goto(url, wait_until="domcontentloaded", timeout=30000)
            except Exception as e:
                print("   не открылась:", e, "\n")
                continue

            # есть ли номер прямо в __NEXT_DATA__ (вдруг с бел. IP не скрыт)?
            try:
                nd = await page.evaluate(
                    "() => document.getElementById('__NEXT_DATA__')?.textContent || ''")
                data = json.loads(nd) if nd else {}
                ad = (data.get("props", {}).get("initialState", {})
                      .get("adView", {}).get("data", {}))
                in_json = [k for k in ad if "phone" in k.lower()]
                print("   ключи *phone* в JSON:", in_json or "нет")
                for k in in_json:
                    print(f"      {k} = {str(ad[k])[:80]}")
            except Exception as e:
                print("   JSON разобрать не вышло:", e)

            # кнопка «Показать телефон»
            clicked = False
            for sel in ["text=Показать телефон", "text=Показать номер",
                        "button:has-text('Показать')", "text=Позвонить"]:
                try:
                    btn = page.locator(sel).first
                    if await btn.count():
                        await btn.scroll_into_view_if_needed(timeout=3000)
                        await btn.click(timeout=4000)
                        clicked = True
                        print(f"   нажал кнопку: {sel}")
                        break
                except Exception:
                    continue
            if not clicked:
                print("   кнопку 'Показать телефон' не нашёл (возможно, номера нет)")

            await page.wait_for_timeout(4000)  # дать reCAPTCHA/запрос отработать

            # номер в DOM (tel: или текстом)
            shown = await page.evaluate(r"""() => {
              const tel = document.querySelector("a[href^='tel:']");
              if (tel) return tel.getAttribute('href').replace('tel:','');
              const m = document.body.innerText.match(/\+375[\s\-()]?\d{2}[\s\-()]?\d{3}[\s\-]?\d{2}[\s\-]?\d{2}/);
              return m ? m[0] : '';
            }""")
            print("   НОМЕР В DOM:", shown or "— не появился —")

            if phone_responses:
                print("   запросы к /phone:")
                for st, u, body in phone_responses:
                    print(f"      [{st}] {u}\n        ответ: {body}")
            else:
                print("   запросов к /phone не зафиксировано")
            print()

        print("Окно закроется через 8 секунд — посмотрите, виден ли номер глазами.")
        await page.wait_for_timeout(8000)
        await browser.close()


if __name__ == "__main__":
    asyncio.run(main())
