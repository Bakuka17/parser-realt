"""
realty_parser_v4.py — Этап 1, версия 4 (debug + lenient parsing).
"""

import asyncio
import hashlib
import re
from datetime import date, timedelta
from pathlib import Path

from playwright.async_api import async_playwright

HAS_STEALTH = False
STEALTH_FN = None
try:
    from playwright_stealth import stealth_async as STEALTH_FN
    HAS_STEALTH = True
except ImportError:
    try:
        from playwright_stealth import Stealth
        _s = Stealth()
        STEALTH_FN = _s.apply_stealth_async
        HAS_STEALTH = True
    except ImportError:
        pass

if not HAS_STEALTH:
    print("⚠ playwright_stealth не установлен — продолжаю без него (это не критично)")

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

UA = ("Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
      "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36")

CATEGORIES = [
    {"url": "https://realt.by/sale/offices/", "deal": "Продажа", "type": "Офис"},
    {"url": "https://realt.by/sale/shops/",   "deal": "Продажа", "type": "Торговое"},
]

LIMIT_PER_CATEGORY = 15
HERE = Path(__file__).parent
OUT_FILE = HERE / "commercial_realty.xlsx"
DEBUG_FILE = HERE / "debug_cards.txt"
DEBUG_DUMP_COUNT = 3

COLUMNS = [
    "Адрес", "Район / Город", "Площадь, м²", "Цена общая", "Цена за м²",
    "Этаж / этажность", "Год постройки", "Класс здания", "Состояние",
    "НДС", "Парковка", "Отдельный вход", "Мокрая зона",
    "Контакт", "Телефон", "Ссылка", "Дата публикации", "Источник",
    "Высота потолков, м", "Грузовая рампа / ворота", "Электр. мощность, кВт",
    "Витринные окна / 1-я линия", "Мин. срок аренды", "Хэш",
]

DEALS = ["Продажа", "Аренда"]
TYPES = ["Офис", "Склад", "Производство", "Торговое", "Общепит", "Здание"]

SECTION_COLORS = {
    "Офис": "4472C4", "Склад": "70AD47", "Производство": "C00000",
    "Торговое": "ED7D31", "Общепит": "7030A0", "Здание": "595959",
}

CITY_MARKERS = [
    "Минск", "Брест", "Витебск", "Гомель", "Гродно", "Могилев", "Могилёв",
    "Барановичи", "Бобруйск", "Борисов", "Молодечно", "Солигорск", "Орша",
    "Пинск", "Полоцк", "Лида", "Новополоцк", "Жлобин", "Светлогорск",
    "Слуцк", "Жодино", "Речица", "Мозырь", "Кобрин", "Рогачёв", "Слоним",
    "Сморгонь", "Волковыск", "Несвиж", "Дзержинск", "Фаниполь",
]
CITY_RE = "|".join(CITY_MARKERS)


def has(pattern, text):
    return "Да" if re.search(pattern, text, re.IGNORECASE) else "н/у"


def parse_features(desc):
    d = (desc or "").lower()
    f = {}
    if re.search(r"с\s*ндс|ндс\s*вкл|вкл\w*\s*ндс", d):
        f["nds"] = "Да"
    elif re.search(r"без\s*ндс|ндс\s*не\s*вкл", d):
        f["nds"] = "Нет"
    else:
        f["nds"] = "н/у"
    f["parking"] = has(r"парковк|паркинг|маши[но\s\-]*мест|стоянк", d)
    f["separate_entrance"] = has(r"отдельн\w*\s*вход", d)
    f["wet_zone"] = has(r"мокр\w*\s*(зон|точк)|санузел|туалет|вода\s+подведен", d)
    m = re.search(r"(?:постройк[аи]|построен\w*|сдан\w*|введен\w*)\s*(?:в\s*)?(\d{4})", d)
    f["year_built"] = m.group(1) if m else "н/у"
    m = re.search(r"класс\w*\s+(prime|a\+?|b\+?|c)\b", d)
    f["building_class"] = m.group(1).upper() if m else "н/у"
    m = re.search(r"высот[ауы]?\s*потолк\w*[\s:]*(\d+[.,]?\d*)\s*м\b", d)
    f["ceiling_height"] = m.group(1).replace(",", ".") if m else "н/у"
    f["ramp_gate"] = has(r"рамп|ворот|грузов\w*\s*въезд", d)
    m = re.search(r"(\d+)\s*к[вВ]т\b", d)
    f["power_kw"] = m.group(1) if m else "н/у"
    f["showcase"] = has(r"витрин\w*\s*окн|перв\w*\s*лини", d)
    if re.search(r"евроремонт", d):
        f["condition"] = "Евроремонт"
    elif re.search(r"после\s*ремонт|с\s*ремонт", d):
        f["condition"] = "После ремонта"
    elif re.search(r"без\s*ремонт|под\s*отделк|строительн\w*\s*вариант", d):
        f["condition"] = "Под отделку"
    else:
        f["condition"] = "н/у"
    return f


def find_address(text):
    lines = [ln.strip() for ln in text.split("\n") if ln.strip()]
    for ln in lines:
        if re.match(rf"г\.\s*({CITY_RE})\b", ln):
            return ln
    for ln in lines:
        if re.search(r"\bр-н\b", ln) and ("ул." in ln or "пр" in ln or "пер" in ln or "просп" in ln or "д." in ln or "," in ln):
            return ln
    for ln in lines:
        if re.match(r"(п\.|д\.|аг\.|пгт)\s*[А-ЯЁ]", ln):
            return ln
    for ln in lines:
        if "с/с" in ln or "Минская область" in ln or "Минский р-н" in ln:
            return ln
    for ln in lines:
        if any(c in ln for c in CITY_MARKERS) and "," in ln and "м²" not in ln:
            return ln
    return ""


def parse_listing_text(text, deal, type_, url):
    t = text.replace("\xa0", " ").replace("\u202f", " ").replace("\u2009", " ")
    price_total = ""
    price_per_m = ""
    m_byn_total = re.search(r"(?<![/\.])(?:от\s+)?([\d\s]{2,})\s*р\.\s*(?![/м])", t)
    if m_byn_total:
        val = re.sub(r"\s+", " ", m_byn_total.group(1)).strip()
        if any(ch.isdigit() for ch in val):
            price_total = val + " р."
    m_byn_perm = re.search(r"(?:от\s+)?([\d\s]{2,})\s*р\.\s*/\s*м²", t)
    if m_byn_perm:
        val = re.sub(r"\s+", " ", m_byn_perm.group(1)).strip()
        if any(ch.isdigit() for ch in val):
            price_per_m = val + " р./м²"
    m_usd_total = re.search(r"≈\s*(?:от\s+)?([\d\s]{2,})\s*\$\s*(?![/м])", t)
    if m_usd_total:
        val = re.sub(r"\s+", " ", m_usd_total.group(1)).strip()
        if any(ch.isdigit() for ch in val):
            price_total = (price_total + " / " if price_total else "") + val + " $"
    m_usd_perm = re.search(r"≈\s*(?:от\s+)?([\d\s]{2,})\s*\$\s*/\s*м²", t)
    if m_usd_perm:
        val = re.sub(r"\s+", " ", m_usd_perm.group(1)).strip()
        if any(ch.isdigit() for ch in val):
            price_per_m = (price_per_m + " / " if price_per_m else "") + val + " $/м²"
    area, floor = "", ""
    m_af = re.search(
        r"(Офис|Магазин|Помещение|Склад|Производ\w*|Кафе|Ресторан|Здание|Бизнес|Торгов\w*|Услуг|Хране)"
        r"([\d.,\- –]+)\s*м²\s*([\d/.\- ]*)(?:\s*этаж)?",
        t,
    )
    if m_af:
        area = m_af.group(2).strip(" -–")
        floor = m_af.group(3).strip()
    address = find_address(t)
    city = ""
    m_city = re.search(rf"г\.\s*({CITY_RE})", address)
    if m_city:
        city = "г. " + m_city.group(1)
    elif re.search(r"([А-ЯЁ][а-яё\-]+)\s+р-н", address):
        m_r = re.search(r"([А-ЯЁ][а-яё\-]+)\s+р-н", address)
        city = m_r.group(1) + " р-н"
    else:
        for cm in CITY_MARKERS:
            if cm in address:
                city = "г. " + cm
                break
    pub_date = ""
    m_d = re.search(r"(\d{1,2}\.\d{1,2}\.\d{4})", t)
    if m_d:
        pub_date = m_d.group(1)
    elif "сегодня" in t.lower():
        pub_date = date.today().strftime("%d.%m.%Y")
    elif "вчера" in t.lower():
        pub_date = (date.today() - timedelta(days=1)).strftime("%d.%m.%Y")
    contact = ""
    if "Агентство" in t:
        contact = "Агентство"
    elif "Контактное лицо" in t:
        contact = "Собственник / Частное лицо"
    desc = t
    if address:
        idx = t.find(address)
        if idx >= 0:
            rest = t[idx + len(address):]
            end_idx = re.search(r"Контакты|Написать|Агентство", rest)
            desc = rest[:end_idx.start()] if end_idx else rest
    feats = parse_features(desc)
    item_hash = hashlib.md5((url + address + area + price_total).encode()).hexdigest()[:12]
    return {
        "Адрес": address, "Район / Город": city, "Площадь, м²": area,
        "Цена общая": price_total, "Цена за м²": price_per_m,
        "Этаж / этажность": floor, "Год постройки": feats["year_built"],
        "Класс здания": feats["building_class"], "Состояние": feats["condition"],
        "НДС": feats["nds"], "Парковка": feats["parking"],
        "Отдельный вход": feats["separate_entrance"], "Мокрая зона": feats["wet_zone"],
        "Контакт": contact, "Телефон": "", "Ссылка": url,
        "Дата публикации": pub_date, "Источник": "realt.by",
        "Высота потолков, м": feats["ceiling_height"],
        "Грузовая рампа / ворота": feats["ramp_gate"],
        "Электр. мощность, кВт": feats["power_kw"],
        "Витринные окна / 1-я линия": feats["showcase"],
        "Мин. срок аренды": "н/у", "Хэш": item_hash,
        "_deal": deal, "_type": type_,
    }


WALK_UP_JS = """
(el) => {
    let cur = el;
    let best = null;
    while (cur && cur.parentElement) {
        const t = (cur.innerText || "").trim();
        if (t.includes("м²") && (t.includes("р.") || t.includes("$"))
            && t.length > 80 && t.length < 4000) {
            best = t;
            if (t.length > 200) break;
        }
        cur = cur.parentElement;
        if (cur && cur.tagName === "BODY") break;
    }
    return best || (el.innerText || "");
}
"""


async def scrape_category(page, cat, limit, debug_dump):
    print(f"\n→ {cat['deal']} / {cat['type']}: {cat['url']}")
    try:
        await page.goto(cat["url"], wait_until="domcontentloaded", timeout=60_000)
    except Exception as e:
        print(f"  ✖ ошибка загрузки: {e}")
        return []
    await asyncio.sleep(6)
    try:
        await page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
        await asyncio.sleep(2)
        await page.evaluate("window.scrollTo(0, 0)")
        await asyncio.sleep(1)
    except Exception:
        pass
    links = await page.query_selector_all("a[href*='/object/']")
    print(f"  ссылок на карточки найдено: {len(links)}")
    results, seen = [], set()
    parsed_count = 0
    empty_address_count = 0
    for link in links:
        if len(results) >= limit:
            break
        href = await link.get_attribute("href")
        if not href:
            continue
        full_url = href if href.startswith("http") else "https://realt.by" + href
        if full_url in seen:
            continue
        seen.add(full_url)
        try:
            text = await link.evaluate(WALK_UP_JS)
        except Exception as e:
            print(f"    ! ошибка извлечения текста: {e}")
            continue
        if not text or len(text) < 50:
            continue
        if len(debug_dump) < DEBUG_DUMP_COUNT:
            debug_dump.append(f"=== CARD #{len(debug_dump) + 1} ({cat['type']}) ===\nURL: {full_url}\nTEXT ({len(text)} симв.):\n{text}\n\n")
        item = parse_listing_text(text, cat["deal"], cat["type"], full_url)
        if not item["Адрес"]:
            empty_address_count += 1
        results.append(item)
        parsed_count += 1
    print(f"  ✓ распарсено объектов: {parsed_count}")
    if empty_address_count:
        print(f"    ⚠ из них с пустым адресом: {empty_address_count}")
    return results


def write_excel(items, output_path):
    wb = Workbook()
    wb.remove(wb.active)
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    section_font = Font(bold=True, color="FFFFFF", size=12)
    thin = Side(style="thin", color="CCCCCC")
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for deal in DEALS:
        ws = wb.create_sheet(deal)
        ws.sheet_view.showGridLines = False
        row = 1
        title_cell = ws.cell(row=row, column=1, value=f"📊 {deal} коммерческой недвижимости")
        title_cell.font = Font(bold=True, size=14)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(COLUMNS))
        row += 2
        for type_ in TYPES:
            type_items = [i for i in items if i["_deal"] == deal and i["_type"] == type_]
            sect = ws.cell(row=row, column=1, value=f"  {type_}  ·  {len(type_items)} объект(ов)")
            sect.font = section_font
            sect.fill = PatternFill("solid", fgColor=SECTION_COLORS[type_])
            sect.alignment = Alignment(vertical="center")
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(COLUMNS))
            ws.row_dimensions[row].height = 22
            row += 1
            for col_idx, name in enumerate(COLUMNS, start=1):
                c = ws.cell(row=row, column=col_idx, value=name)
                c.font = hdr_font
                c.fill = hdr_fill
                c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                c.border = cell_border
            ws.row_dimensions[row].height = 30
            row += 1
            if not type_items:
                ws.cell(row=row, column=1, value="— нет данных в этой выгрузке —").font = Font(italic=True, color="888888")
                row += 2
                continue
            for it in type_items:
                for col_idx, name in enumerate(COLUMNS, start=1):
                    val = it.get(name, "")
                    c = ws.cell(row=row, column=col_idx, value=val)
                    c.alignment = Alignment(vertical="top", wrap_text=True)
                    c.border = cell_border
                    if name == "Ссылка" and val:
                        c.hyperlink = val
                        c.font = Font(color="0563C1", underline="single", size=10)
                    elif name == "Дата публикации":
                        c.alignment = Alignment(horizontal="center", vertical="top")
                    elif val == "н/у":
                        c.font = Font(color="999999", size=10)
                row += 1
            row += 1
        widths = [28, 18, 12, 22, 18, 13, 10, 11, 14, 8, 10, 14, 12, 22, 16, 38, 14, 12, 14, 18, 16, 18, 14, 14]
        for col_idx, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = w
        ws.freeze_panes = "A4"
    summary = wb.create_sheet("Сводка", 0)
    summary.sheet_view.showGridLines = False
    summary["A1"] = "📅 Выгрузка коммерческой недвижимости"
    summary["A1"].font = Font(bold=True, size=16)
    summary["A2"] = f"Дата: {date.today().strftime('%d.%m.%Y')}"
    summary["A2"].font = Font(size=11, color="666666")
    summary["A4"] = "Тип \\ Сделка"
    summary["B4"] = "Продажа"
    summary["C4"] = "Аренда"
    summary["D4"] = "Всего"
    for col in ["A4", "B4", "C4", "D4"]:
        summary[col].font = hdr_font
        summary[col].fill = hdr_fill
        summary[col].alignment = Alignment(horizontal="center", vertical="center")
        summary[col].border = cell_border
    for i, t in enumerate(TYPES, start=5):
        s_count = sum(1 for it in items if it["_deal"] == "Продажа" and it["_type"] == t)
        a_count = sum(1 for it in items if it["_deal"] == "Аренда"  and it["_type"] == t)
        summary[f"A{i}"] = t
        summary[f"A{i}"].fill = PatternFill("solid", fgColor=SECTION_COLORS[t])
        summary[f"A{i}"].font = Font(bold=True, color="FFFFFF")
        summary[f"B{i}"] = s_count
        summary[f"C{i}"] = a_count
        summary[f"D{i}"] = s_count + a_count
        for col in ["A", "B", "C", "D"]:
            summary[f"{col}{i}"].border = cell_border
            summary[f"{col}{i}"].alignment = Alignment(horizontal="center", vertical="center")
    total_row = 5 + len(TYPES)
    summary[f"A{total_row}"] = "ИТОГО"
    summary[f"B{total_row}"] = sum(1 for it in items if it["_deal"] == "Продажа")
    summary[f"C{total_row}"] = sum(1 for it in items if it["_deal"] == "Аренда")
    summary[f"D{total_row}"] = len(items)
    for col in ["A", "B", "C", "D"]:
        summary[f"{col}{total_row}"].font = Font(bold=True, size=12)
        summary[f"{col}{total_row}"].border = cell_border
        summary[f"{col}{total_row}"].alignment = Alignment(horizontal="center", vertical="center")
    summary.column_dimensions["A"].width = 22
    summary.column_dimensions["B"].width = 14
    summary.column_dimensions["C"].width = 14
    summary.column_dimensions["D"].width = 14
    wb.save(output_path)
    print(f"\n✅ Сохранено: {output_path}")
    print(f"   Всего объектов: {len(items)}")


async def main():
    print(f"🚀 Запуск парсера v4. Дата: {date.today():%d.%m.%Y}")
    all_items = []
    debug_dump = []
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=False)
        context = await browser.new_context(user_agent=UA, viewport={"width": 1440, "height": 900}, locale="ru-RU")
        page = await context.new_page()
        if HAS_STEALTH and STEALTH_FN:
            try:
                await STEALTH_FN(page)
            except Exception as e:
                print(f"  ⚠ stealth не применился: {e}")
        for cat in CATEGORIES:
            try:
                items = await scrape_category(page, cat, LIMIT_PER_CATEGORY, debug_dump)
                all_items.extend(items)
            except Exception as e:
                print(f"  ✖ ошибка в категории: {e}")
            await asyncio.sleep(3)
        await browser.close()
    if debug_dump:
        DEBUG_FILE.write_text("\n".join(debug_dump), encoding="utf-8")
        print(f"\n📝 Дебаг сохранён: {DEBUG_FILE}")
        print(f"   В нём — сырой текст первых {len(debug_dump)} карточек. Пришли его мне.")
    write_excel(all_items, OUT_FILE)


if __name__ == "__main__":
    asyncio.run(main())
