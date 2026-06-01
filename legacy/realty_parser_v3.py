"""
realty_parser_v3.py — Прототип Этапа 1.

Что делает:
  • Парсит первую страницу Realt.by по 2 категориям продажи (офисы, торговое).
  • Извлекает ~24 поля по нашей спецификации (часть напрямую, часть — регулярками из описания).
  • Складывает в commercial_realty.xlsx с двумя вкладками (Продажа / Аренда),
    внутри — 6 секций по типу (Офис, Склад, Производство, Торговое, Общепит, Здание).
  • На первой вкладке — сводка (тип × сделка).

Запуск:
  cd ~/realty_env
  source bin/activate          # или: . bin/activate
  pip install openpyxl         # одноразово, если не стоит
  python realty_parser_v3.py

Замечания:
  • Пока работает только Realt.by, только продажа, только 2 категории — это нормально для прототипа.
    После того как Excel-результат тебе понравится — добавим Kufar/Megapolis, аренду и пагинацию.
  • headless=False — намеренно: видно, что грузится; на 2 этапе включим True.
  • Гео-фильтр по умолчанию у Realt — Минск+обл. На 3 этапе обойдём все 6 областей.
"""

import asyncio
import hashlib
import re
from datetime import date, timedelta
from pathlib import Path

from playwright.async_api import async_playwright

try:
    from playwright_stealth import stealth_async
    HAS_STEALTH = True
except ImportError:
    HAS_STEALTH = False
    print("⚠ playwright_stealth не установлен — продолжаю без него")

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ────────────────────────────────────────────────────────────────────
# КОНФИГ
# ────────────────────────────────────────────────────────────────────

UA = (
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
)

CATEGORIES = [
    {"url": "https://realt.by/sale/offices/", "deal": "Продажа", "type": "Офис"},
    {"url": "https://realt.by/sale/shops/",   "deal": "Продажа", "type": "Торговое"},
]

LIMIT_PER_CATEGORY = 15
OUT_FILE = Path(__file__).parent / "commercial_realty.xlsx"

COLUMNS = [
    "Адрес", "Район / Город", "Площадь, м²", "Цена общая", "Цена за м²",
    "Этаж / этажность", "Год постройки", "Класс здания", "Состояние",
    "НДС", "Парковка", "Отдельный вход", "Мокрая зона",
    "Контакт", "Телефон", "Ссылка", "Дата публикации", "Источник",
    "Высота потолков, м", "Грузовая рампа / ворота", "Электр. мощность, кВт",
    "Витринные окна / 1-я линия", "Мин. срок аренды", "Хэш",
]

DEALS = ["Продажа", "Аренда"]
TYPES = ["Офис", "Склад", "Производство", "Торговое", "Общепит", "Здание"]

SECTION_COLORS = {
    "Офис": "4472C4", "Склад": "70AD47", "Производство": "C00000",
    "Торговое": "ED7D31", "Общепит": "7030A0", "Здание": "595959",
}

# ────────────────────────────────────────────────────────────────────
# ПАРСИНГ ОДНОЙ КАРТОЧКИ
# ────────────────────────────────────────────────────────────────────

def has(pattern: str, text: str) -> str:
    return "Да" if re.search(pattern, text, re.IGNORECASE) else "н/у"


def parse_features(desc: str) -> dict:
    d = (desc or "").lower()
    f = {}

    if re.search(r"с\s*ндс|ндс\s*вкл|вкл\w*\s*ндс", d):
        f["nds"] = "Да"
    elif re.search(r"без\s*ндс|ндс\s*не\s*вкл", d):
        f["nds"] = "Нет"
    else:
        f["nds"] = "н/у"

    f["parking"] = has(r"парковк|паркинг|маши[но\s-]*мест|стоянк", d)
    f["separate_entrance"] = has(r"отдельн\w*\s*вход", d)
    f["wet_zone"] = has(r"мокр\w*\s*(зон|точк)|санузел|туалет|вода\s+подведен", d)

    m = re.search(r"(?:постройк[аи]|построен\w*|сдан\w*|введен\w*)\s*(?:в\s*)?(\d{4})", d)
    f["year_built"] = m.group(1) if m else "н/у"

    m = re.search(r"класс\w*\s+(prime|a\+?|b\+?|c)\b", d)
    f["building_class"] = m.group(1).upper() if m else "н/у"

    m = re.search(r"высот[ауы]?\s*потолк\w*[\s:]*(\d+[.,]?\d*)\s*м\b", d)
    f["ceiling_height"] = m.group(1).replace(",", ".") if m else "н/у"

    f["ramp_gate"] = has(r"рамп|ворот|грузов\w*\s*въезд", d)

    m = re.search(r"(\d+)\s*к[вВ]т\b", d)
    f["power_kw"] = m.group(1) if m else "н/у"

    f["showcase"] = has(r"витрин\w*\s*окн|перв\w*\s*лини", d)

    if re.search(r"евроремонт", d):
        f["condition"] = "Евроремонт"
    elif re.search(r"после\s*ремонт|с\s*ремонт", d):
        f["condition"] = "После ремонта"
    elif re.search(r"без\s*ремонт|под\s*отделк|строительн\w*\s*вариант", d):
        f["condition"] = "Под отделку"
    else:
        f["condition"] = "н/у"

    return f


def parse_listing_text(text: str, deal: str, type_: str, url: str) -> dict:
    t = text.replace("\xa0", " ").replace("\u202f", " ")

    # Цены
    price_total = ""
    price_per_m = ""
    for m in re.finditer(r"([\d\s]+)\s*р\.(/м²)?", t):
        val = re.sub(r"\s+", " ", m.group(1)).strip()
        if not val or not any(ch.isdigit() for ch in val):
            continue
        if m.group(2):
            price_per_m = val + " р./м²"
        else:
            price_total = val + " р."
        break  # берём первое вхождение — обычно нужное
    # ищем после р. вариант ≈ XXX $
    m_usd = re.search(r"≈\s*(?:от\s*)?([\d\s]+)\s*\$(/м²)?", t)
    if m_usd:
        val = re.sub(r"\s+", " ", m_usd.group(1)).strip()
        suffix = " / " + val + (" $/м²" if m_usd.group(2) else " $")
        if m_usd.group(2):
            price_per_m = (price_per_m + suffix) if price_per_m else val + " $/м²"
        else:
            price_total = (price_total + suffix) if price_total else val + " $"

    # Площадь и этаж
    area, floor = "", ""
    m_af = re.search(
        r"(Офис|Магазин|Помещение|Склад|Производ\w*|Кафе|Ресторан|Здание|Бизнес|Торгов\w*)"
        r"([\d.,\- –]+)\s*м²([\d/.\- ]*)\s*этаж",
        t,
    )
    if m_af:
        area = m_af.group(2).strip(" -–")
        floor = m_af.group(3).strip()

    # Адрес — пытаемся несколькими паттернами
    address = ""
    for pat in [
        # "г. Минск, ул. Кропоткина, 67" (до перевода строки или ключевого маркера)
        r"(г\.\s*[А-ЯЁ][а-яё\-]+,[^\n]+?)(?=\s*(?:От\s+МКАД|\d+\s*мин|Метро|Контакт|Написать|Агентство|$))",
        # "п. Сонечный, ..., Минский р-н, Минская область"
        r"(п\.\s*[А-ЯЁ][а-яё\-]+,[^\n]+?Минск\w*\s*обл\w*)",
        # "Ждановичский с/с, ..."
        r"([А-ЯЁ][а-яё\-]+\s+с/с,[^\n]+?)(?=\s*(?:Контакт|Написать|$))",
    ]:
        m = re.search(pat, t)
        if m:
            address = m.group(1).strip().rstrip(",")
            break

    # Город / район
    city = ""
    m_city = re.search(r"г\.\s*([А-ЯЁ][а-яё\-]+)", address)
    if m_city:
        city = "г. " + m_city.group(1)
    else:
        m_r = re.search(r"([А-ЯЁ][а-яё\-]+)\s+р-н", address)
        if m_r:
            city = m_r.group(1) + " р-н"

    # Дата
    pub_date = ""
    m_d = re.search(r"(\d{1,2}\.\d{1,2}\.\d{4})", t)
    if m_d:
        pub_date = m_d.group(1)
    elif "сегодня" in t.lower():
        pub_date = date.today().strftime("%d.%m.%Y")
    elif "вчера" in t.lower():
        pub_date = (date.today() - timedelta(days=1)).strftime("%d.%m.%Y")

    # Контакт
    contact = ""
    if "Агентство" in t:
        contact = "Агентство"
    elif "Контактное лицо" in t:
        contact = "Собственник / Частное лицо"

    # Описание (всё после адреса/этажа до "Контакты")
    desc = ""
    m_desc = re.search(r"этаж[\s\S]+?\n(.+?)(?=\bКонтакты\b|\bНаписать\b|\bАгентство\b)", t)
    if m_desc:
        desc = m_desc.group(1).strip()
    else:
        desc = t  # fallback на весь текст карточки

    feats = parse_features(desc)
    item_hash = hashlib.md5((url + address + area + price_total).encode()).hexdigest()[:12]

    return {
        "Адрес": address,
        "Район / Город": city,
        "Площадь, м²": area,
        "Цена общая": price_total,
        "Цена за м²": price_per_m,
        "Этаж / этажность": floor,
        "Год постройки": feats["year_built"],
        "Класс здания": feats["building_class"],
        "Состояние": feats["condition"],
        "НДС": feats["nds"],
        "Парковка": feats["parking"],
        "Отдельный вход": feats["separate_entrance"],
        "Мокрая зона": feats["wet_zone"],
        "Контакт": contact,
        "Телефон": "",  # нет в листинге, только в карточке детали
        "Ссылка": url,
        "Дата публикации": pub_date,
        "Источник": "realt.by",
        "Высота потолков, м": feats["ceiling_height"],
        "Грузовая рампа / ворота": feats["ramp_gate"],
        "Электр. мощность, кВт": feats["power_kw"],
        "Витринные окна / 1-я линия": feats["showcase"],
        "Мин. срок аренды": "н/у",
        "Хэш": item_hash,
        "_deal": deal,
        "_type": type_,
    }


# ────────────────────────────────────────────────────────────────────
# СБОР ДАННЫХ
# ────────────────────────────────────────────────────────────────────

async def scrape_category(page, cat: dict, limit: int) -> list:
    print(f"\n→ {cat['deal']} / {cat['type']}: {cat['url']}")
    try:
        await page.goto(cat["url"], wait_until="domcontentloaded", timeout=60_000)
    except Exception as e:
        print(f"  ✖ ошибка загрузки: {e}")
        return []
    await asyncio.sleep(6)  # время на догрузку JS-листинга

    links = await page.query_selector_all("a[href*='/object/']")
    print(f"  ссылок на карточки найдено: {len(links)}")

    results, seen = [], set()

    for link in links:
        if len(results) >= limit:
            break
        href = await link.get_attribute("href")
        if not href:
            continue
        full_url = href if href.startswith("http") else "https://realt.by" + href
        if full_url in seen:
            continue
        seen.add(full_url)

        try:
            text = await link.evaluate(
                "(el) => { const p = el.closest('article, section, li, div'); "
                "return (p ? p.innerText : el.innerText) || ''; }"
            )
        except Exception:
            text = ""
        if not text or len(text) < 30:
            continue

        item = parse_listing_text(text, cat["deal"], cat["type"], full_url)
        if item["Адрес"]:
            results.append(item)

    print(f"  ✓ распарсено объектов: {len(results)}")
    return results


# ────────────────────────────────────────────────────────────────────
# ВЫГРУЗКА В EXCEL
# ────────────────────────────────────────────────────────────────────

def write_excel(items: list, output_path: Path):
    wb = Workbook()
    wb.remove(wb.active)

    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    section_font = Font(bold=True, color="FFFFFF", size=12)
    thin = Side(style="thin", color="CCCCCC")
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Вкладки сделок ──
    for deal in DEALS:
        ws = wb.create_sheet(deal)
        ws.sheet_view.showGridLines = False
        row = 1

        title_cell = ws.cell(row=row, column=1, value=f"📊 {deal} коммерческой недвижимости")
        title_cell.font = Font(bold=True, size=14)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(COLUMNS))
        row += 2

        for type_ in TYPES:
            type_items = [i for i in items if i["_deal"] == deal and i["_type"] == type_]

            sect = ws.cell(row=row, column=1, value=f"  {type_}  ·  {len(type_items)} объект(ов)")
            sect.font = section_font
            sect.fill = PatternFill("solid", fgColor=SECTION_COLORS[type_])
            sect.alignment = Alignment(vertical="center")
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(COLUMNS))
            ws.row_dimensions[row].height = 22
            row += 1

            for col_idx, name in enumerate(COLUMNS, start=1):
                c = ws.cell(row=row, column=col_idx, value=name)
                c.font = hdr_font
                c.fill = hdr_fill
                c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                c.border = cell_border
            ws.row_dimensions[row].height = 30
            row += 1

            if not type_items:
                ws.cell(row=row, column=1, value="— нет данных в этой выгрузке —").font = Font(italic=True, color="888888")
                row += 2
                continue

            for it in type_items:
                for col_idx, name in enumerate(COLUMNS, start=1):
                    val = it.get(name, "")
                    c = ws.cell(row=row, column=col_idx, value=val)
                    c.alignment = Alignment(vertical="top", wrap_text=True)
                    c.border = cell_border
                    if name == "Ссылка" and val:
                        c.hyperlink = val
                        c.font = Font(color="0563C1", underline="single", size=10)
                    elif name == "Дата публикации":
                        c.alignment = Alignment(horizontal="center", vertical="top")
                    elif val == "н/у":
                        c.font = Font(color="999999", size=10)
                row += 1
            row += 1

        widths = [28, 18, 12, 22, 18, 13, 10, 11, 14, 8, 10, 14, 12, 22, 16, 38, 14, 12, 14, 18, 16, 18, 14, 14]
        for col_idx, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = w
        ws.freeze_panes = "A4"

    # ── Сводка ──
    summary = wb.create_sheet("Сводка", 0)
    summary.sheet_view.showGridLines = False

    summary["A1"] = f"📅 Выгрузка коммерческой недвижимости"
    summary["A1"].font = Font(bold=True, size=16)
    summary["A2"] = f"Дата: {date.today().strftime('%d.%m.%Y')}"
    summary["A2"].font = Font(size=11, color="666666")

    summary["A4"] = "Тип \\ Сделка"
    summary["B4"] = "Продажа"
    summary["C4"] = "Аренда"
    summary["D4"] = "Всего"
    for col in ["A4", "B4", "C4", "D4"]:
        summary[col].font = hdr_font
        summary[col].fill = hdr_fill
        summary[col].alignment = Alignment(horizontal="center", vertical="center")
        summary[col].border = cell_border

    for i, t in enumerate(TYPES, start=5):
        s_count = sum(1 for it in items if it["_deal"] == "Продажа" and it["_type"] == t)
        a_count = sum(1 for it in items if it["_deal"] == "Аренда"  and it["_type"] == t)
        summary[f"A{i}"] = t
        summary[f"A{i}"].fill = PatternFill("solid", fgColor=SECTION_COLORS[t])
        summary[f"A{i}"].font = Font(bold=True, color="FFFFFF")
        summary[f"B{i}"] = s_count
        summary[f"C{i}"] = a_count
        summary[f"D{i}"] = s_count + a_count
        for col in ["A", "B", "C", "D"]:
            summary[f"{col}{i}"].border = cell_border
            summary[f"{col}{i}"].alignment = Alignment(horizontal="center", vertical="center")

    total_row = 5 + len(TYPES)
    summary[f"A{total_row}"] = "ИТОГО"
    summary[f"B{total_row}"] = sum(1 for it in items if it["_deal"] == "Продажа")
    summary[f"C{total_row}"] = sum(1 for it in items if it["_deal"] == "Аренда")
    summary[f"D{total_row}"] = len(items)
    for col in ["A", "B", "C", "D"]:
        summary[f"{col}{total_row}"].font = Font(bold=True, size=12)
        summary[f"{col}{total_row}"].border = cell_border
        summary[f"{col}{total_row}"].alignment = Alignment(horizontal="center", vertical="center")

    summary.column_dimensions["A"].width = 22
    summary.column_dimensions["B"].width = 14
    summary.column_dimensions["C"].width = 14
    summary.column_dimensions["D"].width = 14

    # Источники
    src_row = total_row + 3
    summary[f"A{src_row}"] = "📡 Источники"
    summary[f"A{src_row}"].font = Font(bold=True, size=12)
    src_row += 1
    summary[f"A{src_row}"] = "realt.by"
    summary[f"B{src_row}"] = sum(1 for it in items if it["Источник"] == "realt.by")

    wb.save(output_path)
    print(f"\n✅ Сохранено: {output_path}")
    print(f"   Всего объектов: {len(items)}")


# ────────────────────────────────────────────────────────────────────
# MAIN
# ────────────────────────────────────────────────────────────────────

async def main():
    print(f"🚀 Запуск парсера. Дата: {date.today():%d.%m.%Y}")
    all_items: list = []

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=False)
        context = await browser.new_context(
            user_agent=UA,
            viewport={"width": 1440, "height": 900},
            locale="ru-RU",
        )
        page = await context.new_page()
        if HAS_STEALTH:
            await stealth_async(page)

        for cat in CATEGORIES:
            try:
                items = await scrape_category(page, cat, LIMIT_PER_CATEGORY)
                all_items.extend(items)
            except Exception as e:
                print(f"  ✖ ошибка в категории: {e}")
            await asyncio.sleep(3)

        await browser.close()

    write_excel(all_items, OUT_FILE)


if __name__ == "__main__":
    asyncio.run(main())
