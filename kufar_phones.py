#!/usr/bin/env python3
"""Добор телефонов kufar в commercial_realty.xlsx.

⚠️ ЗАПУСКАТЬ С ВЫКЛЮЧЕННЫМ Psiphon (нужен белорусский IP — иначе kufar прячет номер).

Как работает: берёт kufar-объявления без телефона прямо из xlsx, открывает деталку,
жмёт «Позвонить»/«Показать телефон» (reCAPTCHA v3 проходит сама), ловит ответ
api.kufar.by/.../phone и дописывает номер в колонку «Телефон» той же строки.
Существующие телефоны не трогает. Можно прерывать и запускать снова — продолжит
с оставшихся пустых (естественный резюм).

  ./bin/python kufar_phones.py --limit 25      # первый безопасный прогон
  ./bin/python kufar_phones.py                 # все оставшиеся
  ./bin/python kufar_phones.py --headless      # без видимого окна (рискованнее для капчи)

После добора обновите дашборд: ./bin/python web/export_data.py
"""
import argparse
import asyncio
import json
import random
import re
import shutil
import time
import urllib.request
from pathlib import Path

from openpyxl import load_workbook
from playwright.async_api import async_playwright

import realty_parser_v8 as R

# Маскировка автоматизации (navigator.webdriver и пр.): без неё reCAPTCHA v3 у kufar
# видит робота и молча НЕ отдаёт номер. Тот же стелс, что в realty_parser_v8.
STEALTH_FN = None
try:
    from playwright_stealth import stealth_async as STEALTH_FN
except ImportError:
    try:
        from playwright_stealth import Stealth
        STEALTH_FN = Stealth().apply_stealth_async
    except ImportError:
        pass


async def apply_stealth(page):
    if not STEALTH_FN:
        return
    try:
        await STEALTH_FN(page)
    except Exception as e:
        print(f"  (stealth недоступен: {e})")


def load_chrome_cookies():
    """Вытащить cookies kufar.by из твоего Chrome (где ты залогинен) → формат Playwright.
    kufar спрятал телефон за вход; вход в автоматике невозможен (Google), поэтому берём
    готовую залогиненную сессию из обычного Chrome."""
    import browser_cookie3
    out = []
    for c in browser_cookie3.chrome(domain_name="kufar.by"):
        ck = {"name": c.name, "value": c.value, "domain": c.domain,
              "path": c.path or "/", "secure": bool(c.secure),
              "httpOnly": bool(c.has_nonstandard_attr("HttpOnly")), "sameSite": "Lax"}
        if c.expires:
            ck["expires"] = float(c.expires)
        out.append(ck)
    return out


HERE = Path(__file__).resolve().parent
MAIN_XLSX = HERE / "commercial_realty.xlsx"
PROFILE_DIR = HERE / ".kufar_profile"   # постоянный профиль браузера (залогиненная сессия)
NO_PHONE_CACHE = HERE / ".kufar_nophone.json"  # хэши, у которых телефона реально нет
SHEETS = ("Продажа", "Аренда")
CHECKPOINT_EVERY = 20          # сохранять xlsx каждые N добытых номеров
BAN_STREAK = 12                # столько подряд «кнопка есть, но номер не пришёл» = бан → стоп
PHONE_BTN = ["text=Позвонить", "text=Показать телефон", "text=Показать номер",
             "button:has-text('Показать')"]
USER_AGENTS = [
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 "
    "(KHTML, like Gecko) Version/16.4 Safari/605.1.15",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/124.0 Safari/537.36",
]


def split_phones(raw):
    """Сырьё (может содержать НЕСКОЛЬКО слитых номеров) → список '+375XXXXXXXXX'.

    У объявления бывает 2-3 телефона; kufar отдаёт их подряд. Режем по канону
    бел. номера 375+9 цифр; запасные форматы 80… и общий — одиночные.
    """
    digits = re.sub(r"\D", "", raw or "")
    out = []
    for m in re.findall(r"375\d{9}", digits):       # точная нарезка слипшихся
        p = "+" + m
        if p not in out:
            out.append(p)
    if not out and digits.startswith("80") and len(digits) >= 11:
        out.append("+375" + digits[2:11])
    if not out and digits:
        out.append("+" + digits)
    return out


def norm_phone(raw):
    """Один или несколько номеров → строка через запятую (формат проекта)."""
    return ", ".join(split_phones(raw))


def exit_country():
    """Страна нашего внешнего IP. -> (countryCode, country, ip) или (None,None,None)."""
    try:
        req = urllib.request.Request(
            "http://ip-api.com/json/?fields=country,countryCode,query",
            headers={"User-Agent": "curl/8"})
        with urllib.request.urlopen(req, timeout=8) as r:
            d = json.load(r)
        return d.get("countryCode"), d.get("country", "?"), d.get("query", "?")
    except Exception:
        return None, None, None


def guard_belarus_ip(force):
    """True = можно работать. Под Psiphon (иностранный IP) kufar прячет телефон —
    тратить часы впустую незачем, поэтому при не-белорусском IP останавливаемся."""
    if force:
        print("⚠ Проверка IP пропущена (--force).")
        return True
    cc, country, ip = exit_country()
    if cc is None:
        print("⚠ Не удалось определить страну IP — продолжаю. Если Psiphon включён,"
              " номера НЕ раскроются (тогда прервите и выключите VPN).")
        return True
    if cc != "BY":
        print(f"⛔ Внешний IP — {country} ({ip}), НЕ Беларусь.\n"
              f"   kufar прячет телефоны для иностранных IP. ВЫКЛЮЧИТЕ Psiphon и повторите.\n"
              f"   (обойти проверку, если уверены: --force)")
        return False
    print(f"✓ IP белорусский ({ip}) — kufar отдаст телефоны.")
    return True


def load_nophone():
    try:
        return set(json.loads(NO_PHONE_CACHE.read_text(encoding="utf-8")))
    except Exception:
        return set()


def save_nophone(s):
    try:
        NO_PHONE_CACHE.write_text(json.dumps(sorted(s)), encoding="utf-8")
    except Exception:
        pass


def remaining_count():
    """Сколько kufar ещё можно добрать (без телефона и не в кэше «телефона нет»)."""
    skip = load_nophone()
    wb = load_workbook(MAIN_XLSX, read_only=True)
    n = 0
    for sh in SHEETS:
        if sh not in wb.sheetnames:
            continue
        ws = wb[sh]
        hdr = [c.value for c in ws[2]]
        try:
            ph, sr, hx = hdr.index("Телефон"), hdr.index("Источник"), hdr.index("Хэш")
        except ValueError:
            continue
        for r in ws.iter_rows(min_row=3, values_only=True):
            if len(r) <= max(ph, sr, hx):
                continue
            if "kufar" in str(r[sr] or "") and not str(r[ph] or "").strip():
                if str(r[hx] or "") not in skip:
                    n += 1
    wb.close()
    return n


def collect_targets(ws, limit, skip_hashes):
    """(row, url, ph_col, hash) для kufar-строк без телефона, кроме известных «без тел.».

    Идём С КОНЦА файла вверх: collect_realty дописывает свежие объявления в конец листа
    (base + all_new), значит новые лиды — внизу. Приоритет им: дефицитная дневная квота
    раскрытий kufar тратится сперва на свежие, остаток — на старый хвост (05.07.2026).
    """
    hdr = [c.value for c in ws[2]]
    col = {str(h): i for i, h in enumerate(hdr) if h}
    ph, lk, sr = col.get("Телефон"), col.get("Ссылка"), col.get("Источник")
    hx = col.get("Хэш")
    if ph is None or lk is None or sr is None:
        return []
    out = []
    for r in range(ws.max_row, 2, -1):
        if "kufar" not in str(ws.cell(r, sr + 1).value or ""):
            continue
        if str(ws.cell(r, ph + 1).value or "").strip():
            continue
        h = str(ws.cell(r, hx + 1).value or "") if hx is not None else ""
        if h and h in skip_hashes:          # уже проверяли — телефона нет
            continue
        url = str(ws.cell(r, lk + 1).value or "")
        if url.startswith("http"):
            out.append((r, url, ph + 1, h))
        if limit and len(out) >= limit:
            break
    return out


async def get_phone(page, url):
    """Открыть деталку, раскрыть номер.

    Возвращает (телефон, причина). Причина: ok | no_button (телефона нет, только чат) |
    no_response (кнопка есть, но номер не пришёл) | throttled (kufar ответил 403/429 —
    придушивает). Различение причин нужно, чтобы отличить «реально нет» от «бан».
    """
    caught = []           # номера из ответов /phone
    statuses = []         # коды ответов /phone (детект троттла)

    async def on_resp(resp):
        if "/phone" in resp.url:
            statuses.append(resp.status)
            if resp.status == 200:
                try:
                    j = await resp.json()
                    v = j.get("phone") or j.get("phones")
                    if v:
                        caught.append(v if isinstance(v, str) else ",".join(map(str, v)))
                except Exception:
                    pass

    page.on("response", on_resp)
    reason = "no_button"
    try:
        await page.goto(url, wait_until="domcontentloaded", timeout=30000)
        # «человеческое» поведение — reCAPTCHA v3 оценивает поведение, не галочку
        await page.mouse.wheel(0, random.randint(400, 900))
        await page.wait_for_timeout(random.randint(700, 1600))
        clicked = False
        for sel in PHONE_BTN:
            btn = page.locator(sel).first
            if await btn.count():
                try:
                    await btn.scroll_into_view_if_needed(timeout=2500)
                    await btn.click(timeout=3500)
                    clicked = True
                    break
                except Exception:
                    continue
        if clicked:
            reason = "no_response"
            for _ in range(20):                 # ждём ответ /phone до ~10с
                if caught:
                    break
                await page.wait_for_timeout(500)
            if not caught:                       # fallback: tel:-ссылки в DOM
                dom = await page.evaluate(r"""() => {
                  const tels = [...document.querySelectorAll("a[href^='tel:']")]
                    .map(a => a.getAttribute('href').replace('tel:',''));
                  if (tels.length) return tels.join(',');
                  const m = document.body.innerText.match(/\+375[\s\-()]?\d{2}[\s\-()]?\d{3}[\s\-]?\d{2}[\s\-]?\d{2}/g);
                  return m ? m.join(',') : '';
                }""")
                if dom:
                    caught.append(dom)
        if caught:
            reason = "ok"
        elif clicked and any(s in (403, 429) for s in statuses):
            reason = "throttled"
    finally:
        page.remove_listener("response", on_resp)
    return norm_phone(",".join(caught)), reason


async def do_login():
    """Открыть kufar, дать пользователю войти вручную, сохранить сессию в профиль.
    Пароль вводит ТОЛЬКО пользователь — скрипт лишь открывает окно и ждёт."""
    async with async_playwright() as p:
        ctx = await p.chromium.launch_persistent_context(
            str(PROFILE_DIR), headless=False, locale="ru-RU",
            user_agent=USER_AGENTS[0])
        page = ctx.pages[0] if ctx.pages else await ctx.new_page()
        await apply_stealth(page)
        await page.goto("https://www.kufar.by/account", wait_until="domcontentloaded")
        print("\n→ В открывшемся окне войдите в свой аккаунт kufar (логин/пароль вводите ВЫ).")
        print("  Когда увидите, что вошли, вернитесь сюда и нажмите Enter.")
        await asyncio.get_event_loop().run_in_executor(
            None, input, "  Enter — сохранить сессию: ")
        await ctx.close()
    print(f"✓ Сессия сохранена в {PROFILE_DIR.name}. Теперь запускайте добор обычной командой.")


async def main():
    ap = argparse.ArgumentParser(description="Добор телефонов kufar (бел. IP, VPN OFF).")
    ap.add_argument("--limit", type=int, default=0, help="сколько объявлений за прогон (0 = все)")
    ap.add_argument("--headless", action="store_true", help="без видимого окна (рискованнее)")
    ap.add_argument("--force", action="store_true", help="не проверять страну IP")
    ap.add_argument("--login", action="store_true", help="войти в аккаунт kufar (один раз, руками)")
    ap.add_argument("--chrome-cookies", action="store_true",
                    help="взять логин kufar (cookies) из твоего Chrome — kufar спрятал телефон за вход")
    cfg = ap.parse_args()

    if cfg.login:
        await do_login()
        return

    if not guard_belarus_ip(cfg.force):
        return

    R.acquire_db_lock(HERE / "commercial_realty.xlsx")

    if (HERE / "~$commercial_realty.xlsx").exists():
        print("⚠️  commercial_realty.xlsx открыт в Excel — закройте файл и повторите.")
        return

    print("Бэкап → commercial_realty.xlsx.bak")
    shutil.copy(MAIN_XLSX, MAIN_XLSX.with_suffix(".xlsx.bak"))

    skip = load_nophone()
    wb = load_workbook(MAIN_XLSX)
    targets = []
    per = cfg.limit if cfg.limit else 0
    for sh in SHEETS:
        if sh in wb.sheetnames:
            need = (per - len(targets)) if per else 0
            if per and need <= 0:
                break
            targets += [(sh, *t) for t in collect_targets(wb[sh], need, skip)]
    print(f"К добору: {len(targets)} kufar-объявлений без телефона"
          + (f" (лимит {cfg.limit})" if cfg.limit else "") + "\n")
    if not targets:
        print("Нечего добирать."); return

    got = empty = 0
    async with async_playwright() as p:
        # постоянный профиль: переиспользуем залогиненную сессию (см. --login).
        # Контекст НЕ сбрасываем по ходу — это стёрло бы логин/куки.
        ctx = await p.chromium.launch_persistent_context(
            str(PROFILE_DIR), headless=cfg.headless, locale="ru-RU",
            user_agent=USER_AGENTS[0])
        page = ctx.pages[0] if ctx.pages else await ctx.new_page()
        await apply_stealth(page)
        if cfg.chrome_cookies:
            try:
                cks = load_chrome_cookies()
                await ctx.add_cookies(cks)
                print(f"✓ подставлено {len(cks)} cookies kufar из Chrome (сессия залогинена)")
            except Exception as e:
                print(f"⚠ cookies из Chrome не взялись ({e}) — идём без логина")
        print("режим:", "профиль есть (если входили — залогинен)" if PROFILE_DIR.exists()
              else "без профиля — для большего лимита войдите: --login")
        print()

        from collections import Counter
        reasons = Counter()
        bad_streak = 0
        stop_note = None
        for n, (sheet, row, url, ph_col, h) in enumerate(targets, 1):
            try:
                phone, reason = await get_phone(page, url)
                if not phone and reason == "no_response":   # случайный сбой → 1 повтор
                    await page.wait_for_timeout(random.randint(4000, 8000))
                    phone, reason = await get_phone(page, url)
            except Exception as e:
                phone, reason = "", "error"
                print(f"[{n}/{len(targets)}] {sheet}!{row}  ошибка: {type(e).__name__}")

            reasons[reason] += 1
            if phone:
                wb[sheet].cell(row=row, column=ph_col).value = phone
                got += 1
                bad_streak = 0
                print(f"[{n}/{len(targets)}] {sheet}!{row}  ✓ {phone}")
            else:
                empty += 1
                if reason == "no_button" and h:     # телефона нет → больше не проверяем
                    skip.add(h)
                tag = {"no_button": "нет телефона (только чат)",
                       "no_response": "не отдал номер",
                       "throttled": "⚠ kufar придушивает (403/429)",
                       "error": "ошибка"}.get(reason, reason)
                print(f"[{n}/{len(targets)}] {sheet}!{row}  — {tag}")
                bad_streak = bad_streak + 1 if reason in ("no_response", "throttled") else 0

            if bad_streak >= BAN_STREAK:
                stop_note = (f"{bad_streak} объявлений ПОДРЯД с кнопкой, но без номера — "
                             "kufar придушивает. Останавливаюсь. Сделайте перерыв на "
                             "несколько часов, потом запустите снова — резюм добёрет остаток.")
                break

            if got and got % CHECKPOINT_EVERY == 0:
                wb.save(MAIN_XLSX)
                print(f"   💾 чекпойнт: сохранено (+{got} номеров)")

            await page.wait_for_timeout(random.randint(2500, 6000))  # пауза, как человек

        await ctx.close()

    wb.save(MAIN_XLSX)
    save_nophone(skip)
    if stop_note:
        print(f"\n⛔ {stop_note}")
    print(f"\n📦 Готово: добыто {got}, без номера {empty}.")
    print(f"   разбивка: {dict(reasons)}")
    print("   (no_button = реально только чат; no_response/throttled = kufar не отдал, "
          "добёрётся позже)")
    print(f"   файл: {MAIN_XLSX}")
    print("   Обновите дашборд:  ./bin/python web/export_data.py")


if __name__ == "__main__":
    t0 = time.time()
    asyncio.run(main())
    print(f"   время: {time.time() - t0:.0f}с")
