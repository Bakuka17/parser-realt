#!/usr/bin/env python3
"""Журнал телефонов: единственная копия того, что нельзя пересобрать.

Объявления пересобираются парсером за час. Телефоны kufar — нет: они стоят месяца
ручного добора (белорусский IP, логин в Chrome, дневная квота раскрытий). 09.07.2026
их погибло ~2600, потому что жили ТОЛЬКО в ячейках xlsx.

Отличие от дубля (snapshot_db.py) и от Яндекса: журнал не переписывается НИКОГДА,
только дополняется строкой. Битая запись портит максимум последнюю строку — при
чтении она просто пропускается. Стереть его случайной перезаписью невозможно.

    ./bin/python phones_log.py --dump      # занести в журнал телефоны, что уже в базе
    ./bin/python phones_log.py --restore   # залить номера из журнала в пустые ячейки
    ./bin/python phones_log.py --status    # сколько номеров в журнале и в базе
    ./bin/python phones_log.py --selftest  # самопроверка
"""
import json
import os
from datetime import datetime
from pathlib import Path

HERE = Path(__file__).parent
LOG = HERE / "phones_log.jsonl"
MAIN_XLSX = HERE / "commercial_realty.xlsx"


def _unterminated(path: Path) -> bool:
    """Последняя строка оборвана (нет \\n)? Тогда дозапись склеится с ней и погибнет тоже."""
    if not path.exists() or path.stat().st_size == 0:
        return False
    with path.open("rb") as fh:
        fh.seek(-1, os.SEEK_END)
        return fh.read(1) != b"\n"


def append(hash_: str, phone: str, url: str = "", source: str = "") -> None:
    """Дописать номер. Зовётся СРАЗУ по добыче, до записи в xlsx."""
    if not hash_ or not phone:
        return
    rec = {"ts": f"{datetime.now():%Y-%m-%d %H:%M}", "hash": hash_,
           "phone": phone, "url": url, "source": source}
    prefix = "\n" if _unterminated(LOG) else ""  # закрыть огрызок от прошлого обрыва
    with LOG.open("a", encoding="utf-8") as fh:
        fh.write(prefix + json.dumps(rec, ensure_ascii=False) + "\n")
        fh.flush()
        os.fsync(fh.fileno())  # номер на диске раньше, чем xlsx — переживёт обрыв питания


def load(path: Path | None = None) -> dict[str, str]:
    """{хэш: телефон}. Битые строки пропускаются, поздняя запись побеждает раннюю."""
    path = path or LOG  # не дефолт-аргументом: LOG подменяется в тестах
    out: dict[str, str] = {}
    if not path.exists():
        return out
    for line in path.read_text(encoding="utf-8").splitlines():
        try:
            r = json.loads(line)
            if r.get("hash") and r.get("phone"):
                out[r["hash"]] = r["phone"]
        except json.JSONDecodeError:
            continue  # оборванный хвост последней строки — не повод терять журнал
    return out


def _sheets(wb):
    """(лист, колонка Хэш, колонка Телефон, колонка Ссылка, колонка Источник) по листам с данными."""
    for name in wb.sheetnames:
        ws = wb[name]
        hdr = [c.value for c in ws[2]] if ws.max_row >= 2 else []
        if "Хэш" not in hdr or "Телефон" not in hdr:
            continue
        yield ws, hdr.index("Хэш") + 1, hdr.index("Телефон") + 1, \
            (hdr.index("Ссылка") + 1 if "Ссылка" in hdr else 0), \
            (hdr.index("Источник") + 1 if "Источник" in hdr else 0)


def dump_from_db(xlsx: Path = MAIN_XLSX) -> int:
    """Первичное наполнение: все телефоны, что уже лежат в базе → в журнал."""
    import openpyxl
    known = load()
    wb = openpyxl.load_workbook(xlsx, read_only=True)
    n = 0
    for ws, hc, pc, lc, sc in _sheets(wb):
        for row in ws.iter_rows(min_row=3):
            h = row[hc - 1].value
            p = row[pc - 1].value
            if not h or not p or known.get(str(h)) == str(p):
                continue
            append(str(h), str(p),
                   str(row[lc - 1].value or "") if lc else "",
                   str(row[sc - 1].value or "") if sc else "")
            n += 1
    wb.close()
    return n


def restore_into_db(xlsx: Path = MAIN_XLSX) -> int:
    """Залить номера из журнала в ПУСТЫЕ ячейки «Телефон». Заполненные не трогаем."""
    import openpyxl

    import realty_parser_v8 as R
    phones = load()
    if not phones:
        return 0
    wb = openpyxl.load_workbook(xlsx)
    n = 0
    for ws, hc, pc, _lc, _sc in _sheets(wb):
        for row in range(3, ws.max_row + 1):
            if ws.cell(row=row, column=pc).value:
                continue
            p = phones.get(str(ws.cell(row=row, column=hc).value or ""))
            if p:
                ws.cell(row=row, column=pc).value = p
                n += 1
    if n:
        R.atomic_save(wb, xlsx)
    return n


def _selftest() -> None:
    import tempfile
    global LOG
    tmp = Path(tempfile.mkdtemp())
    LOG = tmp / "log.jsonl"

    append("h1", "375291112233", "https://kufar.by/1", "kufar.by")
    append("h2", "375339998877")
    append("h1", "375291112233, 375445556677")          # уточнение: поздняя побеждает
    append("", "375000000000")                          # без хэша — не пишем
    assert load() == {"h1": "375291112233, 375445556677", "h2": "375339998877"}, load()

    with LOG.open("a", encoding="utf-8") as fh:
        fh.write('{"hash": "h3", "phone": "37529')      # обрыв питания посреди строки
    assert len(load()) == 2, "битый хвост потерял журнал"
    append("h4", "375441234567")                        # журнал продолжает работать
    assert load()["h4"] == "375441234567"
    print("phones_log: 4/4 — дедуп, пропуск без хэша, битый хвост, дозапись после обрыва")


if __name__ == "__main__":
    import argparse
    ap = argparse.ArgumentParser(description="журнал телефонов")
    ap.add_argument("--dump", action="store_true", help="занести телефоны из базы в журнал")
    ap.add_argument("--restore", action="store_true", help="залить журнал в пустые ячейки базы")
    ap.add_argument("--status", action="store_true", help="сколько номеров где")
    ap.add_argument("--selftest", action="store_true")
    a = ap.parse_args()

    if a.selftest:
        _selftest()
    elif a.dump:
        print(f"занесено в журнал: {dump_from_db()} номеров (всего {len(load())})")
    elif a.restore:
        print(f"восстановлено в базу: {restore_into_db()} номеров")
    else:  # --status и пустой вызов
        import openpyxl
        wb = openpyxl.load_workbook(MAIN_XLSX, read_only=True)
        in_db = sum(1 for ws, _h, pc, _l, _s in _sheets(wb)
                    for row in ws.iter_rows(min_row=3) if row[pc - 1].value)
        wb.close()
        print(f"журнал: {len(load())} номеров ({LOG.name})\nбаза:   {in_db} номеров")
