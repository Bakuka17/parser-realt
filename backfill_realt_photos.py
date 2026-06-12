"""Бэкфилл «Фото URL» для realt.by из ЛИСТИНГОВЫХ страниц (SSR __NEXT_DATA__).

Деталки realt фото не отдают (галерея грузится JS-ом, в статике только заглушка),
а ЛИСТИНГИ отдают полный объект с `code` и `images` прямо в HTML (проверено 12.06.2026:
30/30 объектов на странице с фото). Скрипт обходит категории листинга, собирает
карту code → фото и заполняет ПУСТЫЕ «Фото URL» у realt-строк в commercial_realty.xlsx
(формат как у kufar: до 3 URL через «;»). В конце перегенерирует web/data.js.

Объявления, уже снятые с сайта, в листинге не встретятся — их фото не заполнятся,
это ожидаемо (бэкфилл закрывает активные).

Запуск:  ./bin/python backfill_realt_photos.py              # полный проход
         ./bin/python backfill_realt_photos.py --dry-run    # без записи в xlsx
         ./bin/python backfill_realt_photos.py --max-pages 2  # быстрый смоук
"""
import argparse
import json
import random
import re
import subprocess
import sys
import time
import urllib.request
from pathlib import Path

import openpyxl

XLSX = Path(__file__).parent / "commercial_realty.xlsx"
UA = {"User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
      "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120 Safari/537.36"}
CATS = [
    "sale/offices", "sale/shops", "sale/services", "sale/production",
    "sale/restorant-cafe", "sale/warehouses", "sale/storages",
    "rent/offices", "rent/shops", "rent/services", "rent/production",
    "rent/restorant-cafe", "rent/warehouses", "rent/storages",
]
MAX_PHOTOS = 3
NEXT_RE = re.compile(r'<script id="__NEXT_DATA__"[^>]*>(.*?)</script>', re.S)
CODE_RE = re.compile(r'/object/(\d+)')


def fetch(url: str) -> str:
    try:
        req = urllib.request.Request(url, headers=UA)
        return urllib.request.urlopen(req, timeout=30).read().decode("utf-8", "replace")
    except Exception:  # noqa: BLE001 — 404 категории / сетевой сбой = пустая страница
        return ""


def page_objects(html: str) -> dict[str, list[str]]:
    """code → [фото...] со страницы листинга (из __NEXT_DATA__)."""
    m = NEXT_RE.search(html)
    if not m:
        return {}
    try:
        data = json.loads(m.group(1))
    except ValueError:
        return {}
    out: dict[str, list[str]] = {}

    def walk(x):
        if isinstance(x, dict):
            code, imgs = x.get("code"), x.get("images")
            if code and isinstance(imgs, list) and imgs:
                urls = [str(u) for u in imgs if str(u).startswith("http")]
                if urls:
                    out[str(code)] = urls[:MAX_PHOTOS]
            for v in x.values():
                walk(v)
        elif isinstance(x, list):
            for v in x:
                walk(v)

    walk(data.get("props", {}).get("pageProps", {}))
    return out


def crawl(max_pages: int) -> dict[str, list[str]]:
    photo_map: dict[str, list[str]] = {}
    for cat in CATS:
        added_cat, prev_codes = 0, set()
        for page in range(1, max_pages + 1):
            url = f"https://realt.by/{cat}/" + (f"?page={page}" if page > 1 else "")
            objs = page_objects(fetch(url))
            codes = set(objs)
            if not objs or codes == prev_codes:   # конец пагинации / повтор страницы
                break
            prev_codes = codes
            new = {c: u for c, u in objs.items() if c not in photo_map}
            photo_map.update(new)
            added_cat += len(new)
            print(f"  [{cat}] p{page}: объектов {len(objs)}, новых {len(new)}")
            time.sleep(random.uniform(1.0, 2.0))
        print(f"[{cat}] итого новых: {added_cat}")
    return photo_map


def fill_xlsx(photo_map: dict[str, list[str]], dry: bool) -> None:
    if (XLSX.parent / f"~${XLSX.name}").exists():
        sys.exit("⚠ Excel держит файл открытым — закрой его и перезапусти.")
    wb = openpyxl.load_workbook(XLSX)
    total_filled = 0
    for sheet in ("Продажа", "Аренда"):
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        header = [c.value for c in ws[2]]
        col = {name: i + 1 for i, name in enumerate(header)}
        c_link, c_src, c_photo = col["Ссылка"], col["Источник"], col["Фото URL"]
        filled = candidates = 0
        for row in range(3, ws.max_row + 1):
            if ws.cell(row, c_src).value != "realt.by":
                continue
            if ws.cell(row, c_photo).value:        # уже есть фото — не трогаем
                continue
            link = str(ws.cell(row, c_link).value or "")
            m = CODE_RE.search(link)
            if not m:
                continue
            candidates += 1
            urls = photo_map.get(m.group(1))
            if urls:
                ws.cell(row, c_photo).value = ";".join(urls)
                filled += 1
        total_filled += filled
        print(f"[{sheet}] realt-строк без фото: {candidates}, заполнено: {filled}")
    if dry:
        print("(dry-run — xlsx НЕ сохранён)")
        return
    wb.save(XLSX)
    print(f"💾 сохранено: {XLSX.name} (+{total_filled} строк с фото)")
    exporter = XLSX.parent / "web" / "export_data.py"
    if exporter.exists() and total_filled:
        print("→ перегенерирую web/data.js …")
        subprocess.run([sys.executable, str(exporter)], cwd=XLSX.parent)


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--max-pages", type=int, default=80, help="страниц на категорию")
    ap.add_argument("--dry-run", action="store_true", help="не писать в xlsx")
    cfg = ap.parse_args()

    t0 = time.time()
    photo_map = crawl(cfg.max_pages)
    print(f"\n📷 собрано фото-карт: {len(photo_map)} объектов за {time.time()-t0:.0f}с")
    fill_xlsx(photo_map, cfg.dry_run)


if __name__ == "__main__":
    main()
