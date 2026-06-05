"""embed_auctions — встраивает свод аукционов вкладкой «Аукционы» в commercial_realty.xlsx,
чтобы всё было в ОДНОМ файле (не бегать между файлами).

Источник — auctions_realty.xlsx (лист «Аукционы», его делает merge_auctions). Цель —
commercial_realty.xlsx (рядом с листами Сводка/Продажа/Аренда). Существующая вкладка
«Аукционы» в цели пересоздаётся. Вызывается в конце collect_realty и collect_auctions —
так вкладка остаётся свежей при сборе любой из сторон.

Запуск вручную: ./bin/python embed_auctions.py
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HERE = Path(__file__).parent
MAIN = HERE / "commercial_realty.xlsx"
AUCTIONS = HERE / "auctions_realty.xlsx"
SHEET = "Аукционы"

# ширины колонок (как в auctions_common.write_excel)
WIDTHS = {"Сохранить": 10, "Тип торгов": 14, "Тип объекта": 13, "Объект": 40, "Адрес": 28,
          "Район / Город": 16, "Площадь, м²": 12, "Начальная цена": 16, "Задаток": 14,
          "Дата аукциона": 14, "Организатор": 24, "Телефон": 20, "Ссылка": 40,
          "Источник": 16, "Фото URL": 30, "Описание": 70, "Хэш": 14}


def embed(main: Path = MAIN, auctions: Path = AUCTIONS) -> bool:
    """Скопировать лист «Аукционы» из auctions_realty.xlsx в commercial_realty.xlsx. True если сделано."""
    if not main.exists() or not auctions.exists():
        print(f"  (embed_auctions пропущен: нет {main.name if not main.exists() else auctions.name})")
        return False

    src = load_workbook(auctions, read_only=True)
    s = src[SHEET] if SHEET in src.sheetnames else src.active
    rows = [["" if v is None else v for v in r] for r in s.iter_rows(values_only=True)]
    src.close()
    if not rows:
        return False

    dst = load_workbook(main)
    if SHEET in dst.sheetnames:
        del dst[SHEET]
    ws = dst.create_sheet(SHEET)
    for r in rows:
        ws.append(r)

    # строка-шапка = где есть «Ссылка» (под ней — данные)
    hdr_row = next((i for i, r in enumerate(rows, 1) if "Ссылка" in [str(c) for c in r]), 1)
    hdr = [str(c) for c in rows[hdr_row - 1]]
    link_col = hdr.index("Ссылка") + 1 if "Ссылка" in hdr else None

    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill("solid", fgColor="7030A0")
    thin = Side(style="thin", color="CCCCCC")
    cb = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ci, name in enumerate(hdr, 1):
        c = ws.cell(hdr_row, ci)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = cb
        ws.column_dimensions[get_column_letter(ci)].width = WIDTHS.get(name, 14)
    # гиперссылки + перенос в строках данных
    for ri in range(hdr_row + 1, len(rows) + 1):
        for ci in range(1, len(hdr) + 1):
            cell = ws.cell(ri, ci)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if ci == link_col and cell.value:
                cell.hyperlink = cell.value
                cell.font = Font(color="0563C1", underline="single", size=10)
    ws.freeze_panes = ws.cell(hdr_row + 1, 1).coordinate
    ws.auto_filter.ref = f"A{hdr_row}:{get_column_letter(len(hdr))}{len(rows)}"
    ws.sheet_view.showGridLines = False

    dst.save(main)
    print(f"  📑 вкладка «{SHEET}» обновлена в {main.name}: {len(rows) - hdr_row} лотов")
    return True


if __name__ == "__main__":
    embed()
