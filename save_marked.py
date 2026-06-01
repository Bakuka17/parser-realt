"""save_marked — сохраняет отмеченные объекты: фото + гео-анализ локации.

Как пользоваться:
  1. В commercial_realty.xlsx в колонке «Сохранить» поставь любую метку (x, +, 1)
     напротив интересных объектов. Сохрани и ЗАКРОЙ файл.
  2. Запусти: ./bin/python save_marked.py   (или двойной клик по «Сохранить.command»)
  3. Скрипт для каждого отмеченного объекта:
       • скачает фото в photos/{хэш}/ (из колонки «Фото URL»);
       • оценит локацию через бесплатный OpenStreetMap/Overpass:
         «Активность локации» (плотность магазинов/кафе/услуг рядом) + «До транспорта, м»;
       • добавит строку в saved_realty.xlsx (лист «Сохранённые») с путями к фото и гео.
  Уже сохранённые (по хэшу) пропускаются — можно запускать повторно.

Гео без API-ключей (OSM). Рейтинги/отзывы Яндекса и AI-вердикт — отдельно (нужны ключи).
"""
from __future__ import annotations

import json
import math
import time
import urllib.parse
import urllib.request
from datetime import date
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

import realty_parser_v8 as R

HERE = Path(__file__).parent
MAIN_FILE = HERE / "commercial_realty.xlsx"
SAVED_FILE = HERE / "saved_realty.xlsx"
PHOTOS_DIR = HERE / "photos"
UA = "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 Chrome/124 Safari/537.36"
OVERPASS_UA = "realty-tool/1.0 (commercial-realty enrichment)"  # Overpass блокирует браузерные UA
OVERPASS = "https://overpass-api.de/api/interpreter"
NOMINATIM = "https://nominatim.openstreetmap.org/search"
MAX_PHOTOS = 50  # фактически без лимита — для избранного качаем всё что есть в источнике

# доп. колонки листа «Сохранённые» (поверх основных)
EXTRA_COLS = ["Фото (файлы)", "Активность локации", "POI рядом (300м)", "До транспорта, м"]


def is_marked(v) -> bool:
    return v not in (None, "", "н/у") and str(v).strip() != ""


def download_photos(photo_urls: str, hsh: str) -> str:
    """Качает фото в photos/{hsh}/. Возвращает строку с путями через ;."""
    if not photo_urls or photo_urls == "н/у":
        return ""
    dest = PHOTOS_DIR / hsh
    dest.mkdir(parents=True, exist_ok=True)
    saved = []
    for i, url in enumerate(photo_urls.split(";"), 1):
        url = url.strip()
        if not url or i > MAX_PHOTOS:
            continue
        ext = ".jpg"
        for e in (".jpg", ".jpeg", ".png", ".webp"):
            if e in url.lower():
                ext = e
                break
        fp = dest / f"{i}{ext}"
        try:
            req = urllib.request.Request(url, headers={"User-Agent": UA})
            with urllib.request.urlopen(req, timeout=30) as resp:
                data = resp.read()
            if len(data) > 800:  # отсекаем заглушки
                fp.write_bytes(data)
                saved.append(str(fp.relative_to(HERE)))
        except Exception as e:  # noqa: BLE001
            print(f"      ⚠ фото не скачалось: {e}")
        time.sleep(0.3)
    return ";".join(saved)


def geocode(address: str) -> str:
    """Геокодирует адрес через OSM Nominatim. Возвращает 'lat,lng' или ''.
    Бесплатно, без ключа, лимит 1 запрос/сек. Использовать только когда координат нет."""
    if not address or address == "н/у":
        return ""
    try:
        # Добавим «Беларусь» для дис-амбигуирования (адреса часто без страны).
        q = address if "Беларус" in address or "Беларусь" in address else address + ", Беларусь"
        params = urllib.parse.urlencode({"q": q, "format": "json", "limit": "1"})
        req = urllib.request.Request(
            NOMINATIM + "?" + params, headers={"User-Agent": OVERPASS_UA}
        )
        with urllib.request.urlopen(req, timeout=30) as resp:
            data = json.loads(resp.read().decode("utf-8", errors="replace"))
        if data and isinstance(data, list):
            d = data[0]
            return f"{d.get('lat','')},{d.get('lon','')}"
    except Exception as e:  # noqa: BLE001
        print(f"      ⚠ геокодинг адреса не удался: {e}")
    return ""


def _overpass(query: str) -> dict | None:
    try:
        data = urllib.parse.urlencode({"data": query}).encode()
        req = urllib.request.Request(OVERPASS, data=data, headers={"User-Agent": OVERPASS_UA})
        with urllib.request.urlopen(req, timeout=40) as resp:
            return json.loads(resp.read().decode("utf-8", errors="replace"))
    except Exception as e:  # noqa: BLE001
        print(f"      ⚠ Overpass: {e}")
        return None


def _haversine(lat1, lon1, lat2, lon2) -> float:
    r = 6371000
    p1, p2 = math.radians(lat1), math.radians(lat2)
    dp = math.radians(lat2 - lat1)
    dl = math.radians(lon2 - lon1)
    a = math.sin(dp / 2) ** 2 + math.cos(p1) * math.cos(p2) * math.sin(dl / 2) ** 2
    return 2 * r * math.asin(math.sqrt(a))


def geo_enrich(coords: str) -> tuple[str, str, str]:
    """(Активность локации, POI рядом, До транспорта м) по координатам через OSM."""
    if not coords or coords == "н/у" or "," not in coords:
        return "н/у", "н/у", "н/у"
    try:
        lat, lng = (float(x) for x in coords.split(",")[:2])
    except ValueError:
        return "н/у", "н/у", "н/у"

    # 1. Плотность POI в радиусе 300м (магазины/услуги/еда/офисы)
    q_poi = (
        f"[out:json][timeout:30];("
        f"node(around:300,{lat},{lng})[shop];"
        f"node(around:300,{lat},{lng})[amenity~'cafe|restaurant|bank|pharmacy|fast_food|bar|marketplace'];"
        f"node(around:300,{lat},{lng})[office];"
        f");out count;"
    )
    poi = 0
    r1 = _overpass(q_poi)
    if r1 and r1.get("elements"):
        poi = int(r1["elements"][0].get("tags", {}).get("total", 0))
    time.sleep(1.2)

    # 2. Ближайший транспорт в радиусе 800м
    q_tr = (
        f"[out:json][timeout:30];("
        f"node(around:800,{lat},{lng})[highway=bus_stop];"
        f"node(around:800,{lat},{lng})[railway~'station|tram_stop'];"
        f"node(around:800,{lat},{lng})[station=subway];"
        f"node(around:800,{lat},{lng})[public_transport=platform];"
        f");out body;"
    )
    nearest = None
    r2 = _overpass(q_tr)
    if r2 and r2.get("elements"):
        for el in r2["elements"]:
            if "lat" in el and "lon" in el:
                d = _haversine(lat, lng, el["lat"], el["lon"])
                nearest = d if nearest is None else min(nearest, d)
    time.sleep(1.2)

    # классификация активности
    if poi >= 40:
        activity = "высокая"
    elif poi >= 12:
        activity = "средняя"
    elif poi >= 1:
        activity = "низкая"
    else:
        activity = "очень низкая"
    transit = f"{int(nearest)}" if nearest is not None else "нет в 800м"
    return activity, str(poi), transit


def main() -> None:
    import argparse
    p = argparse.ArgumentParser(description="Сохранение отмеченных объектов: фото + гео-анализ.")
    p.add_argument("xlsx", nargs="?", default=str(MAIN_FILE),
                   help="путь к commercial_realty.xlsx")
    p.add_argument("--hashes", default="",
                   help="через запятую: сохранить только эти хэши "
                   "(перекрывает колонку «Сохранить» — нужно для SwiftUI-приложения)")
    ns = p.parse_args()
    main_file = Path(ns.xlsx).expanduser().resolve()
    forced_hashes = {h.strip() for h in ns.hashes.split(",") if h.strip()} if ns.hashes else None

    if not main_file.exists():
        print(f"Нет файла {main_file}")
        return
    print(f"📖 Читаю отметки из {main_file.name}...")
    db, _ = R.load_prev_db(main_file)
    if forced_hashes is not None:
        marked = [it for it in db.values() if str(it.get("Хэш", "")) in forced_hashes]
        print(f"   режим --hashes: выбрано из {len(forced_hashes)} хэшей, найдено {len(marked)}")
    else:
        marked = [it for it in db.values() if is_marked(it.get("Сохранить"))]
    print(f"   отмечено объектов: {len(marked)}")
    if not marked:
        print("   Ничего не отмечено в колонке «Сохранить». Поставь метку (x) и запусти снова.")
        return

    # уже сохранённые хэши
    saved_hashes: set[str] = set()
    if SAVED_FILE.exists():
        try:
            wb0 = load_workbook(SAVED_FILE, read_only=True)
            ws0 = wb0.active
            hdr = [c.value for c in next(ws0.iter_rows(max_row=1))]
            if "Хэш" in hdr:
                hi = hdr.index("Хэш")
                for row in ws0.iter_rows(min_row=2, values_only=True):
                    if hi < len(row) and row[hi]:
                        saved_hashes.add(str(row[hi]))
            wb0.close()
        except Exception as e:  # noqa: BLE001
            print(f"   ⚠ не прочитал saved_realty.xlsx: {e}")

    new = [it for it in marked if str(it.get("Хэш", "")) not in saved_hashes]
    print(f"   новых к сохранению: {len(new)} (уже сохранено: {len(marked) - len(new)})")
    if not new:
        print("   Все отмеченные уже сохранены ранее.")
        return

    records = []
    for i, it in enumerate(new, 1):
        hsh = str(it.get("Хэш", f"row{i}"))
        addr = (it.get("Адрес") or "")[:45]
        print(f"\n[{i}/{len(new)}] {it.get('Тип')} | {addr}")
        photos = download_photos(it.get("Фото URL", ""), hsh)
        print(f"   фото: {len(photos.split(';')) if photos else 0} шт")
        coords = it.get("Координаты", "") or ""
        if not coords or coords == "н/у":
            # координат нет — геокодим адрес через OSM Nominatim
            print("   координат нет → геокодирую адрес…")
            coords = geocode(it.get("Адрес", ""))
            if coords:
                print(f"   ↳ {coords}")
                time.sleep(1.1)  # вежливость к Nominatim
        activity, poi, transit = geo_enrich(coords)
        print(f"   локация: активность={activity} (POI {poi}), транспорт={transit}м")
        rec = dict(it)
        rec["Фото (файлы)"] = photos
        rec["Активность локации"] = activity
        rec["POI рядом (300м)"] = poi
        rec["До транспорта, м"] = transit
        records.append(rec)

    # запись в saved_realty.xlsx (дозапись к существующим).
    # Исключаем «Сохранить» (не нужно в избранных) и «Фото URL» (избыточно — фото
    # уже скачаны, см. колонку «Фото (файлы)»; URL остаётся в основной базе).
    cols = [c for c in R.COLUMNS if c not in ("Сохранить", "Фото URL")] + EXTRA_COLS
    existing_rows = []
    if SAVED_FILE.exists():
        try:
            wbx = load_workbook(SAVED_FILE, read_only=True)
            wsx = wbx.active
            hdrx = [c.value for c in next(wsx.iter_rows(max_row=1))]
            for row in wsx.iter_rows(min_row=2, values_only=True):
                existing_rows.append({hdrx[j]: row[j] for j in range(len(hdrx))})
            wbx.close()
        except Exception:  # noqa: BLE001
            pass

    wb = Workbook()
    ws = wb.active
    ws.title = "Сохранённые"
    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    for ci, name in enumerate(cols, 1):
        c = ws.cell(1, ci, name)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    r = 2
    for rec in existing_rows + records:
        for ci, name in enumerate(cols, 1):
            val = rec.get(name, "")
            cell = ws.cell(r, ci, val)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if name == "Ссылка" and val:
                cell.hyperlink = val
                cell.font = Font(color="0563C1", underline="single")
        r += 1
    ws.freeze_panes = "B2"
    widths = {"Адрес": 30, "Ссылка": 36, "Фото (файлы)": 40, "Фото URL": 36,
              "Активность локации": 16, "До транспорта, м": 15, "Координаты": 18}
    from openpyxl.utils import get_column_letter
    for ci, name in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(ci)].width = widths.get(name, 14)
    wb.save(SAVED_FILE)
    print(f"\n✅ Сохранено в {SAVED_FILE.name}: +{len(records)} (всего {len(existing_rows)+len(records)})")
    print(f"   фото: {PHOTOS_DIR}/")
    # Автоматический бэкап (saved_realty.xlsx + photos/ → iCloud)
    try:
        import collect_realty as CR
        CR.update_memory_and_backup(main_file)
    except Exception as e:  # noqa: BLE001
        print(f"  ⚠ бэкап после сохранения не сработал: {e}")


if __name__ == "__main__":
    main()
