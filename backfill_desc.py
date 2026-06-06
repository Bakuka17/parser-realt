"""backfill_desc — переносит «Описание» из пере-собранных файлов источников
(kufar_realty.xlsx / megapolis_realty.xlsx / realty файлы) в commercial_realty.xlsx по URL.

Зачем: главный файл собран до появления колонки «Описание». Парсеры теперь её извлекают;
этот скрипт добирает описания в существующие строки (аддитивно, строки не трогает, только
пустую «Описание»). Добавляет колонку «Описание», если её ещё нет.

Запуск: ./bin/python backfill_desc.py
"""
from __future__ import annotations

import shutil
from copy import copy
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

import realty_parser_v8 as R

MAIN = Path("commercial_realty.xlsx")
SOURCES = ["kufar_realty.xlsx", "megapolis_realty.xlsx", "realt_desc.xlsx"]


def build_map(path: str) -> dict[str, str]:
    p = Path(path)
    if not p.exists():
        return {}
    m: dict[str, str] = {}
    wb = openpyxl.load_workbook(p, read_only=True)
    for sh in wb.sheetnames:
        ws = wb[sh]
        rows = list(ws.iter_rows(values_only=True))
        hi = next((i for i, r in enumerate(rows) if r and "Ссылка" in [str(c) for c in r]), None)
        if hi is None:
            continue
        hdr = [str(c) for c in rows[hi]]
        if "Описание" not in hdr or "Ссылка" not in hdr:
            continue
        li, di = hdr.index("Ссылка"), hdr.index("Описание")
        for r in rows[hi + 1:]:
            if r[li] and r[di]:
                m[R.normalize_url(str(r[li]))] = r[di]
    wb.close()
    return m


def main() -> None:
    src: dict[str, str] = {}
    for f in SOURCES:
        src.update(build_map(f))
    print(f"описаний в карте источников: {len(src)}")
    if not src or not MAIN.exists():
        print("нечего мержить"); return

    shutil.copy2(MAIN, MAIN.with_suffix(".bak.xlsx"))
    wb = openpyxl.load_workbook(MAIN)
    filled = 0
    for shn in ("Продажа", "Аренда"):
        if shn not in wb.sheetnames:
            continue
        ws = wb[shn]
        hr = next(r for r in ws.iter_rows(min_row=1, max_row=4) if any(str(c.value) == "Ссылка" for c in r))
        hdr = [str(c.value) for c in hr]
        base_row = hr[0].row
        if "Описание" in hdr:
            dcol = hdr.index("Описание") + 1
        else:  # добавить колонку в конец, стиль шапки скопировать с соседней
            dcol = ws.max_column + 1
            hc = ws.cell(base_row, dcol, "Описание")
            sample = hr[-1]
            hc.font = copy(sample.font); hc.fill = copy(sample.fill)
            hc.alignment = copy(sample.alignment); hc.border = copy(sample.border)
            ws.column_dimensions[get_column_letter(dcol)].width = 70
        li = hdr.index("Ссылка") + 1
        for row in ws.iter_rows(min_row=base_row + 1):
            cell = row[dcol - 1]
            if cell.value:
                continue
            u = row[li - 1].value
            key = R.normalize_url(str(u)) if u else None
            if key and key in src:
                cell.value = R.clean_description(src[key])  # применяем актуальную чистку
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                filled += 1
    wb.save(MAIN)
    print(f"✓ заполнено «Описание» в {MAIN.name}: {filled}")


if __name__ == "__main__":
    main()
