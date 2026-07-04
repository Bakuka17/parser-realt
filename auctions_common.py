"""auctions_common — общий каркас для парсеров аукционных площадок РБ.

Все парсеры аукционов (mgcn, e-auction, torgi24, ...) импортируют отсюда:
  • схему колонок AUCTION_COLUMNS;
  • fetch() — загрузка с браузерным UA и ретраями;
  • parse_price() / parse_date() — нормализация (правила цены/дат);
  • make_hash(), write_excel(), load_prev() — дедуп по URL + инкрементал + чекпойнты.

Пишут в auctions_realty.xlsx, лист «Аукционы». Отдельно от commercial_realty.xlsx
(другая природа: торги, а не прямые объявления). Слияние — позже, при необходимости.
"""
from __future__ import annotations

import hashlib
import re
import time
import urllib.parse
import urllib.request
from datetime import date
from pathlib import Path
from typing import Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Side, Border
from openpyxl.utils import get_column_letter

HERE = Path(__file__).parent
DEFAULT_OUT = HERE / "auctions_realty.xlsx"
SHEET = "Аукционы"
UA = ("Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
      "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36")

AUCTION_COLUMNS = [
    "Сохранить",
    "Тип торгов",        # аукцион / конкурс / электронные торги
    "Тип объекта",       # авто из заголовка: Офис/Склад/Земля/... (заполняет write_excel)
    "Объект",            # заголовок лота
    "Адрес",
    "Район / Город",
    "Площадь, м²",
    "Начальная цена",
    "Задаток",
    "Дата аукциона",     # YYYY-MM-DD
    "Организатор",
    "Телефон",
    "Ссылка",
    "Источник",
    "Фото URL",
    "Описание",           # полный текст лота (где доступно; у torgi.gov — из API)
    "Хэш",
]

# ── нормализация цены ────────────────────────────────────────────────────────
_CUR = [
    (re.compile(r"\bBYN\b|бел\.?\s*руб|белорусск\w*\s*рубл|\bBr\b", re.I), "BYN"),
    (re.compile(r"\bUSD\b|\$|долл", re.I), "USD"),
    (re.compile(r"\bEUR\b|€|евро", re.I), "EUR"),
    (re.compile(r"\bRUB\b|рос\.?\s*руб|россий\w*\s*рубл", re.I), "RUB"),
]


def extract_start_price(text: str) -> str:
    """Прицельно вытащить НАЧАЛЬНУЮ цену из текста деталки (схлопывает пробелы/переносы).
    Ловит 'Начальная цена: 945856.58 бел.руб', 'Стартовая цена ... 100 000 BYN'."""
    if not text:
        return ""
    t = re.sub(r"[\s\xa0]+", " ", text.replace("&nbsp;", " "))
    m = re.search(
        r"(?:начальн\w*|стартов\w*)\s+цен\w*[^\d]{0,40}?"
        r"(\d[\d ]*[.,]?\d*)\s*(?:BYN|бел\.?\s*руб|Br|руб)",
        t, re.I,
    )
    if m:
        return parse_price(m.group(0))
    return ""


def parse_price(text: str) -> str:
    """'137 865.03 BYN' → '137865.03 BYN'; '1 250 000 бел. руб.' → '1250000 BYN'."""
    if not text:
        return ""
    cur = ""
    for rx, code in _CUR:
        if rx.search(text):
            cur = code
            break
    # число: цифры, пробелы-разделители тысяч, точка/запятая-десятичная
    m = re.search(r"\d[\d\s]*(?:[.,]\d{1,2})?", text)
    if not m:
        return ""
    num = m.group(0).replace(" ", "").replace(" ", "")
    # запятая как десятичный разделитель → точка
    if "," in num and "." not in num:
        num = num.replace(",", ".")
    else:
        num = num.replace(",", "")
    num = num.rstrip(".")
    if not num or not any(c.isdigit() for c in num):
        return ""
    return f"{num} {cur}".strip()


# ── нормализация дат ─────────────────────────────────────────────────────────
_MONTHS = {
    # русский
    "январ": 1, "феврал": 2, "март": 3, "апрел": 4, "ма": 5, "май": 5, "мая": 5,
    "июн": 6, "июл": 7, "август": 8, "сентябр": 9, "октябр": 10, "ноябр": 11, "декабр": 12,
    # белорусский
    "студзен": 1, "лют": 2, "сакавік": 3, "красавік": 4, "травен": 5, "чэрвен": 6,
    "ліпен": 7, "жнівен": 8, "верасен": 9, "кастрычнік": 10, "лістапад": 11, "снежан": 12,
}


def normalize_price(raw: str) -> "tuple[Optional[float], str]":
    """'151 680.00 BYN' → (151680.0, 'BYN'); '25 000 $' → (25000.0, 'USD'); '999Br' → (999.0,'BYN');
    'договорная'/'по запросу'/''/None → (None, ''). Валюта ∈ {BYN,USD,EUR,RUB,''}.
    Берёт ПЕРВОЕ число и ПЕРВУЮ валюту. Каркас Qwen v2 (17/17 тестов), принят как есть."""
    if raw is None:
        return None, ""
    s = str(raw).strip()
    if not re.search(r"\d", s):
        return None, ""
    m_num = re.search(r"(\d+(?:[\s\xa0]?\d+)*(?:[.,]\d+)?)", s)
    if not m_num:
        return None, ""
    try:
        amount = float(m_num.group(1).replace(" ", "").replace("\xa0", "").replace(",", "."))
    except ValueError:
        return None, ""
    cm = re.search(r"(?i)(BYN|Br|бел\.?\s*руб|белорус.*?руб|USD|\$|долл|EUR|€|евро|"
                   r"RUB|рос\.?\s*руб|россий.*?руб)", s)
    currency = ""
    if cm:
        c = cm.group(1).lower()
        if "byn" in c or c == "br" or "бел" in c:
            currency = "BYN"
        elif "usd" in c or c == "$" or "долл" in c:
            currency = "USD"
        elif "eur" in c or c == "€" or "евро" in c:
            currency = "EUR"
        elif "rub" in c or "рос" in c:
            currency = "RUB"
    return amount, currency


def parse_date(text: str) -> str:
    """'23 июля 2026' / '23.07.2026' / '23 ліпеня 2026' → '2026-07-23'. Иначе ''."""
    if not text:
        return ""
    # DD.MM.YYYY
    m = re.search(r"\b(\d{1,2})[.\-/](\d{1,2})[.\-/](\d{4})\b", text)
    if m:
        d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if 1 <= mo <= 12 and 1 <= d <= 31:
            return f"{y:04d}-{mo:02d}-{d:02d}"
    # DD месяц YYYY (ru/be)
    m = re.search(r"\b(\d{1,2})\s+([а-яёіў]+)\s+(\d{4})", text, re.I)
    if m:
        d = int(m.group(1)); mon_raw = m.group(2).lower(); y = int(m.group(3))
        mo = None
        for stem, num in _MONTHS.items():
            if mon_raw.startswith(stem):
                mo = num
                break
        if mo:
            return f"{y:04d}-{mo:02d}-{int(d):02d}"
    return ""


# ── площадь / адрес (логика от DeepSeek + фикс префикса области) ──────────────
def extract_area(text: str) -> Optional[float]:
    """Площадь в м² из текста: '1156 кв.м'→1156.0, '90,1 кв. м'→90.1, '1 250,5'→1250.5."""
    if not isinstance(text, str):
        return None
    norm = re.sub(r"\s+", " ", text.strip())
    pat = r"(\d[\d\s]*[.,]?\d*)\s*(?:кв\.?\s*м[²2]?|м[²2]|м\.кв\.|S\s*=\s*(\d[\d\s]*[.,]?\d*))"
    m = re.search(pat, norm, re.IGNORECASE)
    if not m:
        m = re.search(r"(?:площад[ья]|S\s*=)\s*(\d[\d\s]*[.,]?\d*)", norm, re.IGNORECASE)
        if not m:
            return None
    num = m.group(1) if m.group(1) else m.group(2)
    if not num:
        return None
    num = re.sub(r"\s", "", num).replace(",", ".")
    try:
        return float(num)
    except ValueError:
        return None


def extract_address(text: str) -> Optional[str]:
    """Адрес из бел./рус. текста торгов. Возвращает строку или None."""
    if not isinstance(text, str):
        return None
    norm = re.sub(r"\s+", " ", text.strip())
    # 1) [Область обл.,] г. Город, ул./пр./пер. Улица, дом  — захват С названием области
    p1 = (r"((?:[А-ЯЁ][а-яё]+\s+обл\.?,?\s+)?"
          r"г\.\s*[А-ЯЁ][а-яё-]+\s*,?\s+"
          r"(?:ул\.|улица|пр\.|пр-т|пр-кт|просп\.|проспект|пер\.|пл\.)\s*"
          r"[А-ЯЁа-яё0-9][А-ЯЁа-яё0-9\-\. ]*?"
          r"\s*,?\s*(?:д\.?\s*)?\d+[А-ЯЁа-яё]*(?:/\d+)?)")
    m = re.search(p1, norm)
    if m:
        return m.group(1).strip().rstrip(",")
    # 2) «в г. Город по ул. Улица, дом» → нормализуем
    m = re.search(r"в\s+г\.\s*([А-ЯЁа-яё-]+)\s+по\s+(ул\.|улица|пр\.|пер\.)\s+([А-ЯЁа-яё\-\. ]+?)\s*,?\s*(\d+[А-ЯЁа-яё]*)", norm, re.IGNORECASE)
    if m:
        return f"г. {m.group(1)}, {m.group(2)} {m.group(3).strip()}, {m.group(4)}"
    # 3) «в городе Город» / «в г. Город» без улицы
    m = re.search(r"(?:в\s+городе|в\s+г\.)\s+([А-ЯЁ][а-яё-]+)", norm)
    if m:
        return m.group(1).strip()
    return None


# ── расширенный извлекатель адреса (каркас DeepSeek, исправлен мной: 10/10) ────
_ADDR_OFFICE = "К.Маркса 39"   # офис MGCN — НИКОГДА не отдавать как адрес объекта
_ADDR_STOP = re.compile(
    r"Наш\s+адрес|Как\s+проехать|Реквизит|Телефон|тел\.|факс|e-mail|организатор|контакт", re.I)
_ADDR_STREET = r"(?:ул\.|улица|пр\.|пр-т|пр-кт|просп\.|проспект|пер\.|пл\.|б-р|бульвар|ш\.|тракт)"
_ADDR_CITY = {"минске": "Минск", "гомеле": "Гомель", "бресте": "Брест", "витебске": "Витебск",
              "могилёве": "Могилёв", "могилеве": "Могилёв", "гродно": "Гродно", "бобруйске": "Бобруйск"}


def _addr_cut(s: str) -> str:
    m = _ADDR_STOP.search(s)
    if m:
        s = s[:m.start()]
    m2 = re.search(r"(?<=[А-Яа-яёЁ]{3})\.\s+[А-ЯЁ]", s)  # граница предложения (не сокращение)
    if m2:
        s = s[:m2.start()]
    return re.sub(r"\s+", " ", s).strip(" .,;")


def extract_re_address(text: str, title: str = "") -> str:
    """Адрес ОБЪЕКТА недвижимости (НЕ офиса/организатора, НЕ юр-болванки). "" если нет.
    Приоритет: маркер «расположен… по адресу:» → шаблон в заголовке «на ул. X, N в Городе»
    → полный адрес в тексте (область/район/НП/улица/дом, с поддержкой «д.NN», «корп. N»)."""
    text = text or ""
    m = re.search(r"располож\w*\s+по\s+адресу:\s*", text, re.I)
    if m:
        a = _addr_cut(text[m.end():m.end() + 170])
        if a and _ADDR_OFFICE not in a:
            return a
    mt = re.search(r"на\s+(" + _ADDR_STREET + r")\s+([^,]+?)\s*,?\s*"
                   r"(\d+[А-Яа-я]?(?:/\d+)?)\s+в\s+([А-Яа-яёЁ]+)", title, re.I)
    if mt:
        c = _ADDR_CITY.get(mt.group(4).lower().rstrip("."), mt.group(4).rstrip("."))
        return f"г. {c}, {mt.group(1)} {mt.group(2).strip()}, {mt.group(3)}"
    body = _ADDR_STOP.split(text)[0]  # отрезаем офис/контакты до поиска
    fp = re.search(r"(?:[А-Я][а-яё]+\s+обл\.?,?\s*)?(?:[А-Я][а-яё-]+\s+р-н,?\s*)?"
                   r"(?:г\.|аг\.|д\.|гп\.)\s*[А-Я][а-яё-]+,?\s*" + _ADDR_STREET +
                   r"\s*[А-Яа-яёЁ0-9\-\.\s]+?,?\s*(?:д\.\s*)?\d+[А-Яа-яёЁ]?(?:/\d+)?"
                   r"(?:\s*,\s*корп\.\s*\d+)?", body, re.I)
    if fp:
        a = re.sub(r"\s+", " ", fp.group(0)).strip(" .,;")
        if a and _ADDR_OFFICE not in a:
            return a
    return ""


# ── классификация типа объекта / номер лота (логика от Qwen + фиксы) ──────────
def classify_object_type(title: str) -> str:
    """Тип объекта по заголовку лота. Один из:
    Офис/Торговое/Склад/Производство/Здание/Квартира/Машиноместо/Земля/Иное."""
    t = (title or "").lower()
    # порядок: квартира ВЫШЕ машиноместа («квартиры с машино-местами» → Квартира)
    if re.search(r"квартир|комнат|жил(?:ое|ой|ая|ые)", t):
        return "Квартира"
    if re.search(r"машино|машиноместо|паркинг|парков", t):
        return "Машиноместо"
    if re.search(r"земельн|право\s+аренды\s+земель|земельны\w*\s+участ|\bучаст(?:ок|ка)", t):
        return "Земля"
    if re.search(r"склад|хранилищ", t):
        return "Склад"
    if re.search(r"производ|цех|\bбаза\b|базы", t):
        return "Производство"
    if re.search(r"торгов|магазин|общепит|кафе|ресторан", t):
        return "Торговое"
    if re.search(r"офис|администрат", t):
        return "Офис"
    if re.search(r"здани|строени|сооружени|комплекс", t):
        return "Здание"
    return "Иное"


def extract_lot_number(title: str) -> Optional[str]:
    """Номер аукциона/лота: '№ 05-А-26' → '05-А-26', 'No 15-А' → '15-А'."""
    m = re.search(r"(?:лот\s*)?(?:№|No|N°)\s*([0-9][0-9А-Яа-яёЁA-Za-z\-]*)", title or "", re.I)
    return m.group(1) if m else None


def extract_area_multi(text: str) -> list[float]:
    """Все площади из КОРОТКОГО текста (заголовок/сниппет, НЕ вся страница!).
    '1156 и 334 кв.м' → [1156.0, 334.0]. На полном HTML не использовать — поймает мусор."""
    if not re.search(r"кв\.?\s*м\.?|м[²2]|площад", text or "", re.I):
        return []
    out = []
    for m in re.findall(r"(\d+(?:\s\d{3})*(?:[,.]\d+)?)(?![А-Яа-яёЁ])", text):
        try:
            out.append(float(m.replace(" ", "").replace(",", ".")))
        except ValueError:
            pass
    return out


# ── телефоны / общее ─────────────────────────────────────────────────────────
PHONE_RE = re.compile(r"\+?375[\s\-()]*\d{2}[\s\-()]*\d{3}[\s\-]*\d{2}[\s\-]*\d{2}")


def _is_placeholder_phone(norm: str) -> bool:
    """Шаблон-заглушка из вёрстки сайта: +375 99 999-99-99, 000-00-00 и т.п.
    Признак — абонентская часть из одной повторяющейся цифры."""
    digits = re.sub(r"\D", "", norm)
    body = digits[3:] if digits.startswith("375") else digits
    if len(body) < 6:
        return True
    return len(set(body)) <= 1


def extract_phones(text: str, limit: int = 3) -> str:
    found = []
    for p in PHONE_RE.findall(text or ""):
        norm = re.sub(r"[^\d+]", "", p)
        if _is_placeholder_phone(norm):
            continue
        if norm not in [re.sub(r"[^\d+]", "", x) for x in found]:
            found.append(p.strip())
        if len(found) >= limit:
            break
    return ", ".join(found)


# ── канон телефонов РБ (каркас Qwen 11/11 + моя защита от фейков; QA на живой базе) ──
_BY_CODE2 = {"17", "25", "29", "33", "44"}                 # Минск-город + мобильные операторы
_BY_CODE3 = {"152", "162", "212", "222", "232"}            # обл. центры (Гродно/Брест/Витебск/Могилёв/Гомель)
_PHONE_CAND = re.compile(r"\+?\d[\d\s()\-]{6,}\d")


def _valid_nat9(d: str) -> bool:
    """9-значный нац. номер с валидным кодом РБ (моб. 2-знач. или гор. 3-знач.)."""
    return len(d) == 9 and (d[:2] in _BY_CODE2 or d[:3] in _BY_CODE3)


def normalize_phones(raw: str) -> list[str]:
    """Все телефоны РБ из строки → канон '+375XXXXXXXXX', дедуп, порядок сохранён.
    Бара 9-значное число принимаем ТОЛЬКО с валидным кодом РБ — иначе цены/УНП/инвентарные
    номера превращались бы в фейк-телефоны (поймано на QA версии Qwen, где правило было len==9)."""
    out, seen = [], set()
    for m in _PHONE_CAND.findall(raw or ""):
        d = re.sub(r"\D", "", m)
        if len(d) == 12 and d.startswith("375"):
            norm = "+" + d
        elif len(d) == 11 and d.startswith("80"):
            norm = "+375" + d[2:] if _valid_nat9(d[2:]) else None
        elif len(d) == 10 and d.startswith("0"):           # формат (017) 360-42-22
            norm = "+375" + d[1:] if _valid_nat9(d[1:]) else None
        elif _valid_nat9(d):
            norm = "+375" + d
        else:
            norm = None
        if norm and norm not in seen:
            seen.add(norm)
            out.append(norm)
    return out


def canon_phones(raw: str) -> str:
    """Поле «Телефон» → канон через запятую; если канона нет — вернуть исходное (не терять данные)."""
    nums = normalize_phones(raw)
    return ", ".join(nums) if nums else (raw or "")


# ── очистка описания лота (каркас от DeepSeek, исправлено и оттестировано мной: 9/9) ──
_DESC_NDS = re.compile(r"Внимание!\s*Цена\s+лота.*?завершения\s+торгов\.\s*", re.I | re.S)
# извещение: до «в HH-MM», затем либо «.» (стиль torgi.gov), либо хвост до
# «Наименование предмет… торгов N» (стиль deloocenka: адрес организатора + шапка таблицы)
_DESC_NOTICE = re.compile(
    r"^.*?\bизвеща(?:ет|ют)\s+о\s+проведении\b.*?\bв\s+\d{1,2}[-:]\d{2}\b"
    r"(?:\s*\.|.*?Наименовани\w*\s+предмет\w*\s+торгов\s*№?\s*\d*\.?)?\s*", re.I | re.S)
_DESC_CONTACTS = re.compile(
    r"Телефон\w*\s+для\s+справок\s*:?.*?"
    r"(?=Кафе|Магазин|Здани|Помещени|Квартир|Дом\b|Коттедж|Гараж|Склад|Офис|Объект|Земельн|"
    r"Капитальн|Сооружени|Строени|Комплекс|Предлага|$)", re.I | re.S)
# телефон РБ: +375 + 9 цифр с любыми разделителями (мобильный 2-знач. код или город 3-знач.)
_DESC_PHONE = re.compile(r"\+?375(?:[\s()\-]*\d){9}")


def clean_auction_description(raw: str) -> str:
    """Сырое описание лота → описание ОБЪЕКТА: срезает юр-преамбулу про НДС
    («Внимание! Цена лота…завершения торгов.»), аукционное извещение организатора
    («ООО X … извещает о проведении … в HH-MM.») и контакты («Телефон для справок: …»,
    номера +375…). Схлопывает пробелы, чистит края (сохраняя финальную точку), режет до 1500."""
    if not raw:
        return ""
    t = _DESC_NDS.sub("", raw)
    t = _DESC_NOTICE.sub("", t)
    t = _DESC_CONTACTS.sub("", t)
    t = _DESC_PHONE.sub("", t)
    t = re.sub(r"\s+", " ", t)
    t = re.sub(r"^[\s.\-–—№]+", "", t).rstrip()
    return t[:1500].rstrip()


def make_hash(url: str, *parts: str) -> str:
    raw = (url.split("?")[0].rstrip("/") + "|" + "|".join(parts)).encode()
    return hashlib.md5(raw).hexdigest()[:12]


def fetch(url: str, retries: int = 3, timeout: int = 30) -> str:
    last: Optional[Exception] = None
    for attempt in range(1, retries + 1):
        try:
            req = urllib.request.Request(
                url, headers={"User-Agent": UA, "Accept-Language": "ru-RU,ru,be"}
            )
            with urllib.request.urlopen(req, timeout=timeout) as resp:
                return resp.read().decode("utf-8", errors="replace")
        except Exception as e:  # noqa: BLE001
            last = e
            time.sleep(1.5 * attempt)
    print(f"    ✖ fetch {url}: {last}")
    return ""


def pdf_text(src, force_ocr: bool = False, lang: str = "rus") -> str:
    """Текст из PDF-извещения (локальный путь ИЛИ URL): берёт текстовый слой,
    а для СКАНОВ без слоя (или force_ocr=True) — OCR (tesseract -l rus).

    ⚠ ПОМЕТКА — ЗВАТЬ ТАМ, ГДЕ НАДО: только для PDF-СКАНОВ. Целевой кейс —
    извещения облисполкомов госимущества (Гомель/Гродно/Брест), которые лежат
    PDF-картинками. Обычные текстовые PDF читаются и так — не тащить OCR в каждый
    парсер ради галочки (YAGNI). OCR медленный (~7с/стр) и для печатного текста
    точный, для декоративных шрифтов/логотипов — мусор.

    Движок ГЛОБАЛЬНЫЙ: ~/.claude/ocr_pdf.py — общий инструмент для ВСЕХ наших
    проектов (правило Дениса), не дублируем логику в репо. Бинарники:
    brew install tesseract poppler (+ rus.traineddata). См. память ocr-for-realty.
    """
    import importlib.util
    import os
    import tempfile

    eng = os.path.expanduser("~/.claude/ocr_pdf.py")
    spec = importlib.util.spec_from_file_location("ocr_pdf", eng)
    if spec is None or not os.path.exists(eng):
        raise RuntimeError(f"OCR-движок не найден: {eng} (поставь, см. память ocr-for-realty)")
    ocr = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(ocr)

    src = str(src)
    if src.startswith(("http://", "https://")):   # PDF по ссылке → временный файл → OCR
        with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
            req = urllib.request.Request(src, headers={"User-Agent": UA})
            with urllib.request.urlopen(req, timeout=60) as r:
                tmp.write(r.read())
            path = tmp.name
        try:
            return ocr.pdf_to_text(path, lang=lang, force_ocr=force_ocr)
        finally:
            os.unlink(path)
    return ocr.pdf_to_text(src, lang=lang, force_ocr=force_ocr)


def clean(s: str) -> str:
    s = re.sub(r"<[^>]+>", " ", s or "")
    import html as _h
    s = _h.unescape(s)
    return re.sub(r"\s+", " ", s).strip()


def get_text(html: str, multiline: bool = False) -> str:
    """HTML → текст: срезает <script>/<style>, убирает теги, схлопывает пробелы.
    multiline=False (деф.): теги→пробел, unescape entity, всё в одну строку.
    multiline=True: теги→\\n, переносы строк сохранены (для парсинга по строкам), без unescape.
    Сведён сюда из 8 копий парсеров (рефактор 24.06) — единый источник."""
    if not html:
        return ""
    import html as _h
    t = re.sub(r"<script.*?</script>", "", html, flags=re.S | re.I)
    t = re.sub(r"<style.*?</style>", "", t, flags=re.S | re.I)
    if multiline:
        t = re.sub(r"<[^>]+>", "\n", t)
        t = re.sub(r"[ \t]+", " ", t)
        return re.sub(r"\n\s*\n", "\n", t).strip()
    t = re.sub(r"<[^>]+>", " ", t)
    return re.sub(r"\s+", " ", _h.unescape(t)).strip()


def blank_item(source: str) -> dict:
    it = {c: "" for c in AUCTION_COLUMNS}
    it["Источник"] = source
    return it


# ── чтение/запись xlsx ───────────────────────────────────────────────────────
def load_prev(path: Path) -> dict[str, dict]:
    """{normalized_url: item} из существующего файла (для инкрементала/дедупа)."""
    db: dict[str, dict] = {}
    if not path.exists():
        return db
    try:
        wb = load_workbook(path, data_only=True, read_only=True)
        if SHEET not in wb.sheetnames:
            wb.close(); return db
        ws = wb[SHEET]
        header = None
        for row in ws.iter_rows(values_only=True):
            if header is None:
                if row and "Ссылка" in row and "Хэш" in row:
                    header = list(row)
                continue
            if not any(v is not None for v in row):
                continue
            rec = {header[i]: row[i] for i in range(len(header)) if header[i]}
            url = rec.get("Ссылка")
            if url:
                # ключ ТОЛЬКО через norm_url — раньше здесь была своя формула
                # (rstrip("/") без unquote) → ключи базы и живого сбора не совпадали,
                # инкремент mgcn пересобирал всё и удваивал файл (поймано 04.07.2026)
                db[norm_url(str(url))] = rec
        wb.close()
    except Exception as e:  # noqa: BLE001
        print(f"  ⚠ не прочитал {path.name}: {e}")
    return db


def norm_url(u: str) -> str:
    # Слеш НЕ срезаем: Bitrix-сайты (ipmtorgi/eauction) без хвостового слеша → 404.
    # unquote: у mgcn один и тот же лот живёт как %d0%ba… И кириллицей → без декода
    # инкремент считал старые лоты новыми и плодил дубли (поймано 04.07.2026).
    u = u or ""
    if "bc.by" in u:  # у bc.by ВСЁ в query (?id=79&news_id=…) — резать нельзя, лоты сливаются в один
        return urllib.parse.unquote(u.split("#")[0])
    return urllib.parse.unquote(u.split("?")[0].split("#")[0])


def write_excel(items: list[dict], path: Path, prev_hashes: Optional[set] = None) -> None:
    """Пишет лист «Аукционы». Новые строки (хэш не в prev_hashes) подсвечивает жёлтым.
    Авто-заполняет «Тип объекта» из «Объект», если пусто (централизованно для всех источников)."""
    prev_hashes = prev_hashes or set()
    for it in items:
        if not it.get("Тип объекта") and it.get("Объект"):
            it["Тип объекта"] = classify_object_type(it["Объект"])
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET
    ws.sheet_view.showGridLines = False
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill("solid", fgColor="7030A0")
    new_fill = PatternFill("solid", fgColor="FFF2A8")
    thin = Side(style="thin", color="CCCCCC")
    cb = Border(left=thin, right=thin, top=thin, bottom=thin)

    c = ws.cell(1, 1, f"🔨 Аукционы недвижимости — {date.today():%d.%m.%Y}")
    c.font = Font(bold=True, size=14)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(AUCTION_COLUMNS))
    for ci, name in enumerate(AUCTION_COLUMNS, 1):
        cc = ws.cell(2, ci, name)
        cc.font = hdr_font; cc.fill = hdr_fill
        cc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cc.border = cb
    # сортировка: по дате аукциона (ближайшие сверху), пустые — вниз
    items_sorted = sorted(items, key=lambda x: (x.get("Дата аукциона") or "9999"))
    row = 3
    new_count = 0
    for it in items_sorted:
        is_new = bool(prev_hashes) and str(it.get("Хэш", "")) not in prev_hashes
        if is_new:
            new_count += 1
        for ci, name in enumerate(AUCTION_COLUMNS, 1):
            val = it.get(name, "")
            if name == "Телефон" and val:
                val = canon_phones(str(val))
            cc = ws.cell(row, ci, val)
            cc.alignment = Alignment(vertical="top", wrap_text=True)
            cc.border = cb
            if name == "Ссылка" and val:
                cc.hyperlink = val
                cc.font = Font(color="0563C1", underline="single", size=10)
            elif name == "Телефон" and val:
                cc.font = Font(bold=True, color="006100", size=11)
            if is_new and name != "Ссылка":
                cc.fill = new_fill
        row += 1
    ws.freeze_panes = "C3"
    widths = {"Сохранить": 10, "Тип торгов": 14, "Тип объекта": 13, "Объект": 40, "Адрес": 28,
              "Район / Город": 16, "Площадь, м²": 12, "Начальная цена": 16,
              "Задаток": 14, "Дата аукциона": 14, "Организатор": 24, "Телефон": 20,
              "Ссылка": 40, "Источник": 16, "Фото URL": 30, "Описание": 70, "Хэш": 14}
    for ci, name in enumerate(AUCTION_COLUMNS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = widths.get(name, 14)
    ws.auto_filter.ref = f"A2:{get_column_letter(len(AUCTION_COLUMNS))}{max(row-1,2)}"
    wb.save(path)
    msg = f"✅ {path.name}: всего {len(items)}"
    if prev_hashes:
        msg += f" | 🆕 новых: {new_count}"
    print(msg)
