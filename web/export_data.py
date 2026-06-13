#!/usr/bin/env python3
"""Экспорт лидов из commercial_realty.xlsx в web/data.js для дашборда обзвона.

Читает листы «Продажа» и «Аренда» (шапка — во 2-й строке, данные с 3-й),
маппит по именам колонок и пишет компактный window.LISTINGS=[...] + window.META.
Запуск: ./bin/python web/export_data.py
"""
import json
import re
from pathlib import Path

from openpyxl import load_workbook

from phones import normalize_phones  # канонизация + дедуп телефонов
from cities import normalize_city    # нормализация населённого пункта

ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT / "commercial_realty.xlsx"
OUT = ROOT / "web" / "data.js"

# имя колонки в xlsx -> ключ в JSON
COLMAP = {
    "Тип": "type",
    "Описание": "desc",
    "Телефон": "phone",
    "Ссылка": "url",
    "Адрес": "addr",
    "Район / Город": "city",
    "Площадь, м²": "area",
    "Цена общая": "price",
    "Цена за м²": "ppm",
    "Дата публикации": "date",
    "Этаж / этажность": "floor",
    "Источник": "source",
    "Фото URL": "photo",
    "Координаты": "coords",
    "Хэш": "hash",
}

# аукционы: своя шапка (AUCTION_COLUMNS) — отдельный маппинг
COLMAP_AUC = {
    "Тип объекта": "type",
    "Объект": "title",
    "Адрес": "addr",
    "Район / Город": "city",
    "Площадь, м²": "area",
    "Начальная цена": "price",
    "Задаток": "deposit",
    "Дата аукциона": "adate",
    "Организатор": "org",
    "Телефон": "phone",
    "Ссылка": "url",
    "Источник": "source",
    "Фото URL": "photo",
    "Описание": "desc",
    "Хэш": "hash",
}

EMPTY = {"", "none", "n/a", "н/у", "нет", "—", "-"}


def clean(v):
    if v is None:
        return ""
    s = str(v).strip()
    return "" if s.lower() in EMPTY else s


def first_photo(s):
    if not s:
        return ""
    for tok in re.split(r"[\s,;]+", s):
        if tok.startswith("http"):
            # kufar: в карточку — миниатюру (~60КБ), а не полноразмер (тяжёлый,
            # через Psiphon грузится наполовину). Полные фото качает «Сохранить».
            return tok.replace("/v1/gallery/", "/v1/list_thumbs_2x/")
    return ""


def usd(price_str):
    """Достаём $-число из строки вида '303 600 р. / 109 849 $'."""
    if not price_str:
        return None
    m = re.search(r"([\d][\d   ]*)\s*\$", price_str)
    if not m:
        return None
    digits = re.sub(r"[   ]", "", m.group(1))
    try:
        v = int(digits)
    except ValueError:
        return None
    return v if v > 0 else None  # '0 $' = цена скрыта, считаем как отсутствует


def to_float(s):
    if not s:
        return None
    m = re.search(r"[\d]+(?:[.,]\d+)?", s)
    if not m:
        return None
    try:
        return float(m.group(0).replace(",", "."))
    except ValueError:
        return None


# Город нормализуется общим normalize_city из cities.py (импорт выше):
# срезает тип поселения (г./аг./д./п./с.с.), схлопывает дубли (г. Аг. Колодищи → Колодищи).


def parse_coords(s):
    if not s:
        return None
    m = re.match(r"\s*(-?\d+\.\d+)\s*,\s*(-?\d+\.\d+)\s*", s)
    if not m:
        return None
    return [float(m.group(1)), float(m.group(2))]


def load_sheet(ws, deal, sheet_name):
    rows = ws.iter_rows(values_only=True)
    next(rows, None)            # строка 1 — заголовок-плашка
    header = next(rows, None)   # строка 2 — настоящая шапка
    if not header:
        return []
    idx = {}
    for i, name in enumerate(header):
        name = (str(name).strip() if name is not None else "")
        if name in COLMAP:
            idx[COLMAP[name]] = i
    out = []
    # данные начинаются с 3-й строки Excel; enumerate держит реальный номер строки
    for excel_row, r in enumerate(rows, start=3):
        if not r or all(c is None for c in r):
            continue
        rec = {k: clean(r[i]) if i < len(r) else "" for k, i in idx.items()}
        if not rec.get("hash") and not rec.get("url"):
            continue
        price = rec.get("price", "")
        if not re.search(r"[1-9]", price):  # только нули/пусто → цена скрыта
            price = ""
        item = {
            "deal": deal,
            "sheet": sheet_name,
            "row": excel_row,
            "type": rec.get("type", ""),
            "phone": ", ".join(normalize_phones(rec.get("phone", ""))),
            "url": rec.get("url", ""),
            "addr": rec.get("addr", ""),
            "city": normalize_city(rec.get("city", "")),
            "area": to_float(rec.get("area", "")),
            "price": price,
            "usd": usd(price),
            "date": rec.get("date", ""),
            "floor": rec.get("floor", ""),
            "source": rec.get("source", ""),
            "photo": first_photo(rec.get("photo", "")),
            "coords": parse_coords(rec.get("coords", "")),
            "desc": (rec.get("desc", "")[:200]),
            "hash": rec.get("hash", ""),
        }
        out.append(item)
    return out


def parse_adate(s):
    """Дата аукциона → ('dd.mm.yyyy', future?). future=None если даты нет."""
    s = s or ""
    m = re.search(r"(\d{4})-(\d{2})-(\d{2})", s)
    if m:
        y, mo, d = m.group(1), m.group(2), m.group(3)
    else:
        m = re.search(r"(\d{2})\.(\d{2})\.(\d{4})", s)
        if not m:
            return "", None
        d, mo, y = m.group(1), m.group(2), m.group(3)
    import datetime as _dt
    try:
        dt = _dt.date(int(y), int(mo), int(d))
    except ValueError:
        return "", None
    return f"{int(d):02d}.{int(mo):02d}.{y}", dt >= _dt.date.today()


def load_auctions(ws):
    rows = ws.iter_rows(values_only=True)
    next(rows, None)            # строка 1 — плашка
    header = next(rows, None)   # строка 2 — шапка
    if not header:
        return []
    idx = {}
    for i, name in enumerate(header):
        name = str(name).strip() if name is not None else ""
        if name in COLMAP_AUC:
            idx[COLMAP_AUC[name]] = i
    out = []
    for excel_row, r in enumerate(rows, start=3):
        if not r or all(c is None for c in r):
            continue
        rec = {k: clean(r[i]) if i < len(r) else "" for k, i in idx.items()}
        if not rec.get("hash") and not rec.get("url"):
            continue
        adate, future = parse_adate(rec.get("adate", ""))
        out.append({
            "deal": "auction",
            "sheet": "Аукционы",
            "row": excel_row,
            "type": rec.get("type", "") or "Аукцион",
            "title": rec.get("title", ""),
            "phone": ", ".join(normalize_phones(rec.get("phone", ""))),
            "url": rec.get("url", ""),
            "addr": rec.get("addr", ""),
            "city": normalize_city(rec.get("city", "")),
            "area": to_float(rec.get("area", "")),
            "price": rec.get("price", ""),        # «Начальная цена» (BYN, бывает диапазон)
            "deposit": rec.get("deposit", ""),    # задаток
            "date": adate,                        # дата аукциона (dd.mm.yyyy)
            "future": future,                     # True/False/None — актуален ли
            "org": rec.get("org", ""),
            "source": rec.get("source", ""),
            "photo": first_photo(rec.get("photo", "")),
            "coords": None,
            "desc": (rec.get("desc", "")[:200]),
            "hash": rec.get("hash", ""),
        })
    return out


def main():
    wb = load_workbook(SRC, read_only=True)
    items = []
    for sheet, deal in (("Продажа", "sale"), ("Аренда", "rent")):
        if sheet in wb.sheetnames:
            items += load_sheet(wb[sheet], deal, sheet)
    if "Аукционы" in wb.sheetnames:
        items += load_auctions(wb["Аукционы"])
    wb.close()

    meta = {
        "total": len(items),
        "withPhone": sum(1 for x in items if x["phone"]),
        "withPhoto": sum(1 for x in items if x["photo"]),
        "withCoords": sum(1 for x in items if x["coords"]),
        "sale": sum(1 for x in items if x["deal"] == "sale"),
        "rent": sum(1 for x in items if x["deal"] == "rent"),
        "auction": sum(1 for x in items if x["deal"] == "auction"),
        "auctionFuture": sum(1 for x in items if x["deal"] == "auction" and x.get("future")),
        "generated": __import__("datetime").datetime.now().strftime("%d.%m.%Y %H:%M"),
    }

    OUT.parent.mkdir(exist_ok=True)
    payload = json.dumps(items, ensure_ascii=False, separators=(",", ":"))
    metaj = json.dumps(meta, ensure_ascii=False)
    OUT.write_text(
        f"window.META={metaj};\nwindow.LISTINGS={payload};\n", encoding="utf-8"
    )
    size_mb = OUT.stat().st_size / 1e6
    print(f"OK: {len(items)} объектов → {OUT} ({size_mb:.1f} MB)")
    print("META:", metaj)


if __name__ == "__main__":
    main()
