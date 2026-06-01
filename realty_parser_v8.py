"""realty_parser_v8 — realt.by: список + детальные страницы (телефоны), выгрузка в Excel."""
from __future__ import annotations

import argparse
import asyncio
import hashlib
import random
import re
from dataclasses import dataclass
from datetime import date, timedelta
from pathlib import Path
from typing import Any, Callable, Optional, Set

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from playwright.async_api import Page, async_playwright

HAS_STEALTH = False
STEALTH_FN: Optional[Callable[..., Any]] = None
try:
    from playwright_stealth import stealth_async as STEALTH_FN

    HAS_STEALTH = True
except ImportError:
    try:
        from playwright_stealth import Stealth

        _s = Stealth()
        STEALTH_FN = _s.apply_stealth_async
        HAS_STEALTH = True
    except ImportError:
        pass

UA = (
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
)

CATEGORIES = [
    {"url": "https://realt.by/sale/offices/", "deal": "Продажа", "type": "Офис"},
    {"url": "https://realt.by/sale/shops/", "deal": "Продажа", "type": "Торговое"},
    {"url": "https://realt.by/sale/services/", "deal": "Продажа", "type": "Торговое"},
    {"url": "https://realt.by/sale/warehouses/", "deal": "Продажа", "type": "Склад"},
    {"url": "https://realt.by/sale/storages/", "deal": "Продажа", "type": "Склад"},
    {"url": "https://realt.by/sale/production/", "deal": "Продажа", "type": "Производство"},
    {"url": "https://realt.by/sale/restorant-cafe/", "deal": "Продажа", "type": "Общепит"},
    {"url": "https://realt.by/rent/offices/", "deal": "Аренда", "type": "Офис"},
    {"url": "https://realt.by/rent/shops/", "deal": "Аренда", "type": "Торговое"},
    {"url": "https://realt.by/rent/services/", "deal": "Аренда", "type": "Торговое"},
    {"url": "https://realt.by/rent/warehouses/", "deal": "Аренда", "type": "Склад"},
    {"url": "https://realt.by/rent/storages/", "deal": "Аренда", "type": "Склад"},
    {"url": "https://realt.by/rent/production/", "deal": "Аренда", "type": "Производство"},
    {"url": "https://realt.by/rent/restorant-cafe/", "deal": "Аренда", "type": "Общепит"},
]

DEFAULT_LIMIT_PER_CATEGORY = 10
DEFAULT_FETCH_DETAILS = True
DEFAULT_HEADLESS = True
CHECKPOINT_EVERY = 50  # сохранять xlsx каждые N новых карточек (защита от потери данных)
HERE = Path(__file__).parent
DEFAULT_OUT_FILE = HERE / "commercial_realty.xlsx"
DEFAULT_DEBUG_FILE = HERE / "debug_cards.txt"

COLUMNS = [
    "Сохранить",
    "Тип",
    "Телефон",
    "Ссылка",
    "Адрес",
    "Район / Город",
    "Площадь, м²",
    "Цена общая",
    "Цена за м²",
    "НДС",
    "Этаж / этажность",
    "Год постройки",
    "Класс здания",
    "Состояние",
    "Парковка",
    "Отдельный вход",
    "Мокрая зона",
    "Контакт",
    "Имя контакта",
    "Дата публикации",
    "Источник",
    "Высота потолков, м",
    "Грузовая рампа / ворота",
    "Электр. мощность, кВт",
    "Витринные окна / 1-я линия",
    "Мин. срок аренды",
    "Материал стен",
    "Фото URL",
    "Координаты",
    "Хэш",
]
DEALS = ["Продажа", "Аренда"]
TYPE_ORDER = ["Офис", "Склад", "Производство", "Торговое", "Общепит", "Здание"]
TYPE_COLORS = {
    "Офис": "4472C4",
    "Склад": "70AD47",
    "Производство": "C00000",
    "Торговое": "ED7D31",
    "Общепит": "7030A0",
    "Здание": "595959",
}
CITY_MARKERS = [
    "Минск",
    "Брест",
    "Витебск",
    "Гомель",
    "Гродно",
    "Могилев",
    "Могилёв",
    "Барановичи",
    "Бобруйск",
    "Борисов",
    "Молодечно",
    "Солигорск",
    "Орша",
    "Пинск",
    "Полоцк",
    "Лида",
    "Новополоцк",
    "Жлобин",
    "Светлогорск",
    "Слуцк",
    "Жодино",
    "Речица",
    "Мозырь",
    "Кобрин",
    "Рогачёв",
    "Слоним",
    "Сморгонь",
    "Волковыск",
    "Несвиж",
    "Дзержинск",
    "Фаниполь",
]
CITY_RE = "|".join(CITY_MARKERS)

PHONE_BY = re.compile(r"\+?375\s*\(?\d{2}\)?\s*\d{3}[-\s]?\d{2}[-\s]?\d{2}")
PHONE_FALLBACK = re.compile(
    r"\+?\d{1,3}[\s\-]?\(?\d{2,4}\)?[\s\-]?\d{3}[\s\-]?\d{2}[\s\-]?\d{2}"
)


@dataclass
class RunConfig:
    limit_per_category: int
    fetch_details: bool
    headless: bool
    out_file: Path
    debug_file: Optional[Path]
    goto_retries: int
    verbose: bool
    max_pages: int = 100
    full_rescan: bool = False  # игнорировать БД и перепрогнать всё


def normalize_url(u: str) -> str:
    return u.split("?")[0].split("#")[0].rstrip("/")


def has(p: str, t: str) -> str:
    return "Да" if re.search(p, t, re.IGNORECASE) else "н/у"


def normalize_nds(value: str) -> str:
    v = (value or "").strip().lower()
    if not v:
        return "н/у"
    if re.search(r"ндс\s*нет|ндс\s*0\s*%|без\s*ндс|ндс\s*не\s*включ|не\s*вкл\.?\s*ндс", v):
        return "Нет"
    if re.search(r"с\s*ндс|ндс\s*вкл|включая\s*ндс|\+?\s*ндс", v):
        return "Да"
    if v in {"да", "есть"}:
        return "Да"
    if v in {"нет", "не", "отсутствует"}:
        return "Нет"
    return "н/у"


DETAIL_VALUE_BLOCKLIST = frozenset(
    {
        "Год постройки", "Материал стен", "Ремонт", "Электрическая мощность",
        "Площадь участка", "Площадь общая", "Принадлежность объекта",
        "Район города", "Микрорайон", "НДС", "С НДС", "Цена с НДС",
        "Класс", "Высота потолков", "Парковка", "Отдельный вход", "Санузел",
        "Отопление", "Электроснабжение", "Естественное освещение", "Вода",
        "Расположение", "Юридический адрес", "Номер договора", "Примечание",
        "Местоположение", "Область", "Населенный пункт", "Улица", "Номер дома",
        "Этаж", "Этажность", "Этаж / этажность", "Тип объекта", "Тип",
        "Раздельных помещений", "Мебель", "Интернет", "Компьютерная сеть",
        "Сигнализация", "Видеонаблюдение", "Скидка", "Подробнее",
        "Показать больше", "Показать контакты",
    }
)


def normalize_yesno(value: str) -> str:
    v = (value or "").strip()
    if not v or v in DETAIL_VALUE_BLOCKLIST:
        return "н/у"
    vl = v.lower()
    if vl in {"есть", "да", "имеется"}:
        return "Да"
    if vl in {"нет", "отсутствует", "—", "-"}:
        return "Нет"
    if re.search(r"\d+\s*санузл|санузел|туалет", vl):
        return "Да"
    if re.search(r"открыт|подземн|гостев|крыт|маши[но\s\-]*мест", vl):
        return "Да"
    return v


BUILDING_PATTERNS = [
    r"отдельно\s*стоящ\w*\s+здани",
    r"отдельн\w*\s+здани",
    r"здани\w*\s+целиком",
    r"целиком\s+здани",
    r"административн\w*\s+здани",
    r"административно[\-\s]\w+\s+здани",
    r"производственн\w*\s+здани",
    r"имуществен\w*\s+комплекс",
    r"капитальн\w*\s+строени\w*\s+целиком",
    r"^\s*здание\b",  # заголовок начинается со слова «Здание»
]
_BUILDING_RE = re.compile("|".join(BUILDING_PATTERNS), re.IGNORECASE | re.MULTILINE)


def is_building_text(text: str) -> bool:
    """Эвристика: похоже ли объявление на продажу/аренду здания целиком.
    Консервативно — отсекаем «помещение в здании» (там здание не объект)."""
    if not text:
        return False
    low = text.lower()
    # явный негатив: единица внутри здания
    if re.search(r"помещени\w*\s+в\s+здани|в\s+бизнес[\-\s]?центре|в\s+бц\b", low):
        # но если рядом есть «отдельно стоящее» — всё равно здание
        if not re.search(r"отдельно\s*стоящ", low):
            return False
    return bool(_BUILDING_RE.search(text))


def parse_features(ft: str, d: str) -> dict:
    ft = (ft or "").lower()
    d = (d or "").lower()
    f: dict = {}
    f["nds"] = (
        "Да"
        if re.search(r"с\s*ндс|ндс\s*вкл", d)
        else ("Нет" if re.search(r"без\s*ндс", d) else "н/у")
    )
    f["parking"] = has(r"парковк|паркинг|маши[но\s\-]*мест|стоянк", d)
    f["separate_entrance"] = has(r"отдельн\w*\s*вход", d)
    f["wet_zone"] = has(r"мокр\w*\s*(зон|точк)|санузел|туалет", d)
    m = re.search(r"класс\w*\s+(prime|a\+?|b\+?|c)\b", ft)
    f["building_class"] = m.group(1).upper() if m else "н/у"
    m = re.search(r"высот[ауы]?\s*потолк\w*[\s:]*(\d+[.,]?\d*)\s*м\b", d)
    f["ceiling_height"] = m.group(1).replace(",", ".") if m else "н/у"
    f["ramp_gate"] = has(r"рамп|ворот|грузов\w*\s*въезд", d)
    f["showcase"] = has(r"витрин\w*\s*окн|перв\w*\s*лини", d)
    m = re.search(r"(?:срок\s+аренды|мин\w*\s+срок)[\s\.:]*([\d]+)\s*мес", d)
    f["min_rent"] = (
        m.group(1) + " мес" if m else ("долгосрочно" if "долгосрочн" in d else "н/у")
    )
    return f


def find_address(text: str) -> str:
    lines = [ln.strip() for ln in text.split("\n") if ln.strip()]
    for ln in lines:
        if re.match(rf"г\.\s*({CITY_RE})\b", ln):
            return ln
    for ln in lines:
        if re.search(r"\bр-н\b", ln) and (
            "ул." in ln or "пр" in ln or "пер" in ln or "просп" in ln or "," in ln
        ):
            return ln
    for ln in lines:
        if re.match(r"(п\.|д\.|аг\.|пгт)\s*[А-ЯЁ]", ln):
            return ln
    for ln in lines:
        if "с/с" in ln or "Минская область" in ln:
            return ln
    for ln in lines:
        if any(c in ln for c in CITY_MARKERS) and "," in ln and "м²" not in ln:
            return ln
    return ""


def find_area_floor(text: str) -> tuple[str, str]:
    lines = [ln.strip() for ln in text.split("\n") if ln.strip()]
    area, floor = "", ""
    for ln in lines:
        if not area:
            m = re.match(r"^([\d.,\- –]+)\s*м²\s*$", ln)
            if m:
                area = m.group(1).strip(" -–")
                continue
        if not floor:
            m = re.match(r"^([\d/.\-]+)\s*этаж\s*$", ln)
            if m:
                floor = m.group(1).strip()
        if area and floor:
            break
    return area, floor


def find_date(text: str) -> str:
    # «Контакты» в подвале («Контакты редакции») даёт ложные срабатывания — сначала блок объявления.
    for marker in ("Контактное лицо", "Размещено", "Обновлено", "Опубликовано", "Контакты"):
        idx = text.find(marker)
        if idx >= 0:
            region = text[idx : idx + 500]
            m = re.search(r"(\d{1,2})\.(\d{1,2})\.(\d{4})", region)
            if m:
                return f"{int(m.group(1)):02d}.{int(m.group(2)):02d}.{m.group(3)}"
            low = region.lower()
            if "сегодня" in low:
                return date.today().strftime("%d.%m.%Y")
            if "вчера" in low:
                return (date.today() - timedelta(days=1)).strftime("%d.%m.%Y")
    m = re.search(r"(\d{1,2})\.(\d{1,2})\.(\d{4})", text)
    if m:
        return f"{int(m.group(1)):02d}.{int(m.group(2)):02d}.{m.group(3)}"
    low = text.lower()
    if "сегодня" in low:
        return date.today().strftime("%d.%m.%Y")
    if "вчера" in low:
        return (date.today() - timedelta(days=1)).strftime("%d.%m.%Y")
    return ""


def _digit_count(s: str) -> int:
    return sum(1 for c in s if c.isdigit())


def extract_phones(text: str, tel_hrefs: Optional[list[str]] = None) -> list[str]:
    if tel_hrefs:
        from_tel = []
        for href in tel_hrefs:
            digits = re.sub(r"\D", "", href)
            if digits.startswith("375") and len(digits) >= 12:
                from_tel.append("+" + digits)
        if from_tel:
            return from_tel
    phones = PHONE_BY.findall(text)
    if phones:
        return phones
    raw = PHONE_FALLBACK.findall(text)
    return [p for p in raw if 9 <= _digit_count(p) <= 15]


def parse_listing_text(text: str, deal: str, type_: str, url: str) -> dict:
    t = text.replace("\xa0", " ").replace("\u202f", " ").replace("\u2009", " ")
    price_total = ""
    price_per_m = ""
    m = re.search(r"(?<![/\.])(?:от\s+)?([\d\s]{2,})\s*р\.\s*(?![/м])", t)
    if m:
        v = re.sub(r"\s+", " ", m.group(1)).strip()
        if any(c.isdigit() for c in v):
            price_total = v + " р."
    m = re.search(r"(?:от\s+)?([\d\s]{2,})\s*р\.\s*/\s*м²", t)
    if m:
        v = re.sub(r"\s+", " ", m.group(1)).strip()
        if any(c.isdigit() for c in v):
            price_per_m = v + " р./м²"
    m = re.search(r"≈\s*(?:от\s+)?([\d\s]{2,})\s*\$\s*(?![/м])", t)
    if m:
        v = re.sub(r"\s+", " ", m.group(1)).strip()
        if any(c.isdigit() for c in v):
            price_total = (price_total + " / " if price_total else "") + v + " $"
    m = re.search(r"≈\s*(?:от\s+)?([\d\s]{2,})\s*\$\s*/\s*м²", t)
    if m:
        v = re.sub(r"\s+", " ", m.group(1)).strip()
        if any(c.isdigit() for c in v):
            price_per_m = (price_per_m + " / " if price_per_m else "") + v + " $/м²"
    area, floor = find_area_floor(t)
    address = find_address(t)
    city = ""
    city_m = re.search(rf"г\.\s*({CITY_RE})", address)
    district_m = re.search(r"([А-ЯЁ][а-яё\-]+)\s+р-н", address)
    if city_m:
        city = "г. " + city_m.group(1)
    elif district_m:
        city = district_m.group(1) + " р-н"
    else:
        for cm in CITY_MARKERS:
            if cm in address:
                city = "г. " + cm
                break
    pub_date = find_date(t)
    contact = ""
    if "Агентство" in t:
        contact = "Агентство"
    elif "Контактное лицо" in t:
        contact = "Собственник / Частное лицо"
    desc = t
    if address:
        i = t.find(address)
        if i >= 0:
            rest = t[i + len(address) :]
            e = re.search(r"Контакты|Написать|Агентство", rest)
            desc = rest[: e.start()] if e else rest
    feats = parse_features(t, desc)
    final_type = "Здание" if is_building_text(t) else type_
    h = hashlib.md5(
        (normalize_url(url) + address + area + price_total).encode()
    ).hexdigest()[:12]
    return {
        "Сохранить": "",
        "Тип": final_type,
        "Адрес": address,
        "Район / Город": city,
        "Площадь, м²": area,
        "Цена общая": price_total,
        "Цена за м²": price_per_m,
        "Этаж / этажность": floor,
        "Год постройки": "н/у",
        "Класс здания": feats["building_class"],
        "Состояние": "н/у",
        "НДС": feats["nds"],
        "Парковка": feats["parking"],
        "Отдельный вход": feats["separate_entrance"],
        "Мокрая зона": feats["wet_zone"],
        "Контакт": contact,
        "Имя контакта": "",
        "Телефон": "",
        "Ссылка": normalize_url(url),
        "Дата публикации": pub_date,
        "Источник": "realt.by",
        "Высота потолков, м": feats["ceiling_height"],
        "Грузовая рампа / ворота": feats["ramp_gate"],
        "Электр. мощность, кВт": "н/у",
        "Витринные окна / 1-я линия": feats["showcase"],
        "Мин. срок аренды": feats["min_rent"],
        "Материал стен": "н/у",
        "Фото URL": "",
        "Координаты": "",
        "Хэш": h,
        "_deal": deal,
    }


WALK_UP_JS = """
(el) => {
    let cur = el; let best = null;
    while (cur && cur.parentElement) {
        const t = (cur.innerText || "").trim();
        if (t.includes("м²") && (t.includes("р.") || t.includes("$"))
            && t.length > 80 && t.length < 4000) {
            best = t;
            if (t.length > 200) break;
        }
        cur = cur.parentElement;
        if (cur && cur.tagName === "BODY") break;
    }
    return best || (el.innerText || "");
}
"""


def debug_append(cfg: RunConfig, line: str) -> None:
    if not cfg.debug_file:
        return
    try:
        with cfg.debug_file.open("a", encoding="utf-8") as f:
            f.write(line.rstrip() + "\n")
    except OSError:
        pass


async def apply_stealth(page: Page, cfg: RunConfig) -> None:
    if not (HAS_STEALTH and STEALTH_FN):
        return
    try:
        await STEALTH_FN(page)
    except Exception as e:
        if cfg.verbose:
            print(f"  (stealth) {e}")


async def goto_with_retry(
    page: Page,
    url: str,
    *,
    cfg: RunConfig,
    wait_until: str = "domcontentloaded",
    timeout: int = 60_000,
) -> bool:
    last: Optional[Exception] = None
    for attempt in range(1, cfg.goto_retries + 1):
        try:
            await page.goto(url, wait_until=wait_until, timeout=timeout)
            return True
        except Exception as e:
            last = e
            if cfg.verbose:
                print(f"    goto {attempt}/{cfg.goto_retries}: {e}")
            await asyncio.sleep(1.2 * attempt)
    if cfg.verbose and last:
        print(f"    goto окончательно не удалось: {last}")
    return False


async def fetch_details(page: Page, item: dict, cfg: RunConfig) -> dict:
    """Карточка: параметры с страницы, клик «Показать контакты», телефоны."""
    url = item["Ссылка"]
    if not await goto_with_retry(page, url, cfg=cfg, timeout=30_000):
        debug_append(cfg, f"FAIL goto\t{url}")
        return item
    await asyncio.sleep(random.uniform(2, 4))
    try:
        full = await page.evaluate("() => document.body.innerText || ''")
    except Exception as e:
        if cfg.verbose:
            print(f"    evaluate body: {e}")
        full = ""
    # ФОТО REALT — НЕ ИЗВЛЕКАЕМ.
    # Эксперимент показал: единственный openximg-URL, видимый в <img> после загрузки,
    # одинаковый для всех карточек (общая og:image-заглушка). Реальные фото галереи
    # подгружаются позже / лежат в CSS background-image / в JS-компоненте — нужен
    # глубокий реверс. Пока оставляем «Фото URL» пустым у realt; гео-анализ для
    # realt-строк работает через геокодинг адреса в save_marked.py (OSM Nominatim).
    # Координаты у realt подгружаются JS — попробуем найти в DOM/тексте; если нет,
    # save_marked.py геокодит адрес через OSM Nominatim в момент сохранения.
    try:
        coords_str = await page.evaluate(
            """() => {
                const txt = document.body.innerText || '';
                // ymaps.init({ center: [53.9, 27.5] }) или JSON с lat/lng
                let m = txt.match(/(\\d{2}\\.\\d{4,8})[,\\s]+(\\d{2}\\.\\d{4,8})/);
                if (m) return m[1] + ',' + m[2];
                const meta = document.querySelector('[data-coord], [data-coords], [data-latlng]');
                return meta ? (meta.getAttribute('data-coord') || meta.getAttribute('data-coords') || meta.getAttribute('data-latlng')) : '';
            }"""
        )
        if coords_str and "," in coords_str:
            item["Координаты"] = coords_str
    except Exception:
        pass
    lines = [ln.strip() for ln in full.split("\n") if ln.strip()]
    params: dict[str, str] = {}
    detail_labels = {
        "Год постройки",
        "Материал стен",
        "Ремонт",
        "Электрическая мощность",
        "Площадь участка",
        "Принадлежность объекта",
        "Район города",
        "Микрорайон",
        "НДС",
        "С НДС",
        "Цена с НДС",
        "Класс",
        "Высота потолков",
        "Парковка",
        "Отдельный вход",
        "Санузел",
        "Тип объекта",
    }
    for i, ln in enumerate(lines):
        if i + 1 < len(lines) and ln in detail_labels:
            nx = lines[i + 1]
            if nx not in DETAIL_VALUE_BLOCKLIST:
                params[ln] = nx
    if "Год постройки" in params:
        m = re.search(r"\d{4}", params["Год постройки"])
        if m:
            item["Год постройки"] = m.group(0)
    if "Материал стен" in params:
        item["Материал стен"] = params["Материал стен"]
    if "Ремонт" in params:
        item["Состояние"] = params["Ремонт"]
    if "Электрическая мощность" in params:
        m = re.search(r"(\d+)\s*кВт", params["Электрическая мощность"])
        if m:
            item["Электр. мощность, кВт"] = m.group(1)
    if "Класс" in params:
        cls = params["Класс"].strip().upper()
        if re.fullmatch(r"(PRIME|A\+?|B\+?|C)", cls):
            item["Класс здания"] = cls
    if "Высота потолков" in params:
        m = re.search(r"(\d+[.,]?\d*)", params["Высота потолков"])
        if m:
            item["Высота потолков, м"] = m.group(1).replace(",", ".")
    if "Парковка" in params:
        v = normalize_yesno(params["Парковка"])
        if v != "н/у":
            item["Парковка"] = v
    if "Отдельный вход" in params:
        v = normalize_yesno(params["Отдельный вход"])
        if v != "н/у":
            item["Отдельный вход"] = v
    if "Санузел" in params:
        v = normalize_yesno(params["Санузел"])
        if v != "н/у":
            item["Мокрая зона"] = v
    # Точная классификация «Здание» по структурному полю «Тип объекта».
    if "Тип объекта" in params:
        if re.search(r"здани|отдельно\s*стоящ|строени\w*\s+целиком", params["Тип объекта"], re.I):
            item["Тип"] = "Здание"
    elif is_building_text(full):
        item["Тип"] = "Здание"
    nds_candidates = [params.get("НДС"), params.get("С НДС"), params.get("Цена с НДС")]
    for nds_value in nds_candidates:
        normalized = normalize_nds(nds_value or "")
        if normalized != "н/у":
            item["НДС"] = normalized
            break
    # Иногда телефон уже присутствует в tel:-ссылках без клика.
    try:
        pre_text, pre_tel_hrefs = await page.evaluate(
            """() => [
                document.body.innerText || '',
                [...document.querySelectorAll("a[href^='tel:']")].map(a => a.getAttribute('href'))
            ]"""
        )
        pre_phones = extract_phones(pre_text, pre_tel_hrefs)
        if pre_phones:
            item["Телефон"] = ", ".join(sorted(set(pre_phones[:3])))
            return item
    except Exception:
        pass
    try:
        contact_btns = page.locator(
            "button:has-text('Показать контакты'), "
            "a:has-text('Показать контакты'), "
            "button:has-text('Показать контакт')"
        )
        btn_count = await contact_btns.count()
        clicked = False
        for idx in range(min(btn_count, 5)):
            btn = contact_btns.nth(idx)
            try:
                if not await btn.is_visible():
                    continue
                await btn.scroll_into_view_if_needed(timeout=3000)
                await btn.click(timeout=7000)
                clicked = True
                break
            except Exception:
                continue
        if not clicked and btn_count:
            try:
                await contact_btns.first.click(timeout=5000, force=True)
                clicked = True
            except Exception:
                pass
        if clicked:
            await asyncio.sleep(random.uniform(0.8, 1.6))
    except Exception as e:
        if cfg.verbose:
            print(f"    клик контакты: {type(e).__name__}")
    try:
        new_text, tel_hrefs = await page.evaluate(
            """() => [
                document.body.innerText || '',
                [...document.querySelectorAll("a[href^='tel:']")].map(a => a.getAttribute('href'))
            ]"""
        )
        phones = extract_phones(new_text, tel_hrefs)
        if phones:
            item["Телефон"] = ", ".join(sorted(set(phones[:3])))
        name_m = re.search(
            r"(?:Продавец|Контактное лицо|Собственник|Контакт)\s*\n+"
            r"([А-ЯЁ][а-яёА-ЯЁ\- ]{1,40}?)\s*\n",
            new_text,
        )
        if name_m:
            candidate = name_m.group(1).strip()
            if not re.search(r"\d|Показать|контакт", candidate, re.IGNORECASE):
                item["Имя контакта"] = candidate
        if not phones and cfg.debug_file:
            debug_append(cfg, f"NO_PHONE\t{url}\tchars={len(new_text)}")
    except Exception as e:
        if cfg.verbose:
            print(f"    телефоны: {e}")
    return item


def parse_pub_date(s: str) -> Optional[date]:
    """Парсит '04.05.2026' → date. Возвращает None если не получилось."""
    if not s:
        return None
    m = re.search(r"(\d{1,2})\.(\d{1,2})\.(\d{4})", s)
    if not m:
        return None
    try:
        return date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
    except ValueError:
        return None


async def scrape_one_page(
    page: Page, page_url: str, cat: dict, global_seen: Set[str], cfg: RunConfig
) -> list[dict]:
    """Парсит одну страницу листинга. Возвращает список item-dict (без деталей)."""
    if not await goto_with_retry(page, page_url, cfg=cfg, timeout=60_000):
        print(f"  ✖ не удалось открыть {page_url}")
        return []
    await asyncio.sleep(random.uniform(3, 5))
    try:
        await page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
        await asyncio.sleep(1.5)
        await page.evaluate("window.scrollTo(0, 0)")
        await asyncio.sleep(0.8)
    except Exception as e:
        if cfg.verbose:
            print(f"  scroll: {e}")
    links = await page.query_selector_all("a[href*='/object/']")
    items: list[dict] = []
    for link in links:
        href = await link.get_attribute("href")
        if not href:
            continue
        full_url = href if href.startswith("http") else "https://realt.by" + href
        norm = normalize_url(full_url)
        if norm in global_seen:
            continue
        global_seen.add(norm)
        try:
            text = await link.evaluate(WALK_UP_JS)
        except Exception as e:
            if cfg.verbose:
                print(f"  walk_up skip: {e}")
            continue
        if not text or len(text) < 50:
            if cfg.debug_file:
                debug_append(cfg, f"SHORT_CARD\t{norm}\tlen={len(text or '')}")
            continue
        items.append(parse_listing_text(text, cat["deal"], cat["type"], full_url))
    return items


async def scrape_category(
    page: Page,
    cat: dict,
    limit: int,
    global_seen: Set[str],
    cfg: RunConfig,
    prev_urls: Optional[Set[str]] = None,
    last_run_date: Optional[date] = None,
) -> list:
    """Идёт по страницам категории. Стоп-условия:
       • достигнут лимит limit (если задан, обычно очень большой);
       • страница не дала ни одного нового URL (конец пагинации);
       • вся страница состоит из уже известных URL (incremental mode);
       • все даты публикации на странице < last_run_date (incremental mode).
    """
    print(f"\n→ {cat['deal']} / {cat['type']}: {cat['url']}")
    prev_urls = prev_urls or set()
    results: list[dict] = []
    for page_n in range(1, cfg.max_pages + 1):
        sep = "&" if "?" in cat["url"] else "?"
        page_url = cat["url"] if page_n == 1 else f"{cat['url']}{sep}page={page_n}"
        items = await scrape_one_page(page, page_url, cat, global_seen, cfg)
        if not items:
            if cfg.verbose:
                print(f"  стр.{page_n}: 0 новых ссылок → стоп")
            break
        # категоризируем по знакомым / новым
        new_items = [it for it in items if normalize_url(it["Ссылка"]) not in prev_urls]
        known_count = len(items) - len(new_items)
        results.extend(items)
        # инкрементальные стоп-условия
        all_known = bool(prev_urls) and not new_items
        # дата-катофф: все валидные даты на странице старше last_run_date
        date_stop = False
        if last_run_date and not cfg.full_rescan:
            dates = [parse_pub_date(it.get("Дата публикации", "")) for it in items]
            valid = [d for d in dates if d is not None]
            if valid and all(d < last_run_date for d in valid):
                date_stop = True
        print(
            f"  стр.{page_n}: всего {len(items)} | новых {len(new_items)} | "
            f"известных {known_count}"
            + (" | DATE_STOP" if date_stop else "")
            + (" | ALL_KNOWN" if all_known else "")
        )
        if len(results) >= limit:
            print(f"  ✋ достигнут лимит {limit}")
            break
        if all_known and not cfg.full_rescan:
            break
        if date_stop:
            break
        if page_n < cfg.max_pages:
            await asyncio.sleep(random.uniform(1.5, 3))
    print(f"  ✓ из категории всего: {len(results)}")
    return results


def load_prev_hashes(path: Path) -> dict[str, set[str]]:
    """Возвращает {deal: {hash, ...}} из предыдущей выгрузки (если файл есть)."""
    result: dict[str, set[str]] = {}
    if not path.exists():
        return result
    try:
        from openpyxl import load_workbook

        wb = load_workbook(path, data_only=True, read_only=True)
        for sname in wb.sheetnames:
            if sname not in DEALS:
                continue
            ws = wb[sname]
            rows = ws.iter_rows(values_only=True)
            header = None
            hashes: set[str] = set()
            for row in rows:
                if header is None:
                    if row and "Хэш" in (row or ()):
                        header = list(row)
                    continue
                if not header:
                    continue
                try:
                    hi = header.index("Хэш")
                except ValueError:
                    break
                if hi < len(row) and row[hi]:
                    hashes.add(str(row[hi]))
            if hashes:
                result[sname] = hashes
        wb.close()
    except Exception as e:
        print(f"  ⚠ не смог прочитать предыдущую выгрузку ({e})")
    return result


def load_prev_db(path: Path) -> tuple[dict[str, dict], Optional[date]]:
    """Возвращает (db: {normalized_url: item_dict}, last_run_date).

    item_dict содержит все колонки + ключ '_deal' (Продажа/Аренда), как в свежем парсинге.
    last_run_date берётся из Сводки (строка 'Дата: DD.MM.YYYY').
    """
    db: dict[str, dict] = {}
    last_run: Optional[date] = None
    if not path.exists():
        return db, last_run
    try:
        from openpyxl import load_workbook

        wb = load_workbook(path, data_only=True, read_only=True)
        # last_run_date из Сводки
        if "Сводка" in wb.sheetnames:
            s = wb["Сводка"]
            for row in s.iter_rows(min_row=1, max_row=8, values_only=True):
                for cell in row or ():
                    if isinstance(cell, str) and cell.startswith("Дата:"):
                        last_run = parse_pub_date(cell)
                        break
                if last_run:
                    break
        for sname in wb.sheetnames:
            if sname not in DEALS:
                continue
            ws = wb[sname]
            header: Optional[list] = None
            for row in ws.iter_rows(values_only=True):
                if header is None:
                    if row and "Ссылка" in (row or ()) and "Хэш" in (row or ()):
                        header = list(row)
                    continue
                if not any(v is not None for v in row):
                    continue
                rec = {h: row[i] for i, h in enumerate(header) if h}
                url = rec.get("Ссылка")
                if not url:
                    continue
                rec["_deal"] = sname
                db[normalize_url(str(url))] = rec
        wb.close()
    except Exception as e:
        print(f"  ⚠ не смог прочитать БД из {path}: {e}")
    return db, last_run


def write_excel(
    items: list,
    output_path: Path,
    prev_hashes: Optional[dict[str, set[str]]] = None,
) -> None:
    # prev_hashes можно передать снаружи (для чекпойнтов), иначе читаем из файла.
    if prev_hashes is None:
        prev_hashes = load_prev_hashes(output_path)
    new_fill = PatternFill("solid", fgColor="FFF2A8")  # мягкий жёлтый для новых строк
    wb = Workbook()
    wb.remove(wb.active)
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    thin = Side(style="thin", color="CCCCCC")
    cb = Border(left=thin, right=thin, top=thin, bottom=thin)
    type_idx = {t: i for i, t in enumerate(TYPE_ORDER)}
    new_counts: dict[str, int] = {d: 0 for d in DEALS}
    for deal in DEALS:
        ws = wb.create_sheet(deal)
        ws.sheet_view.showGridLines = False
        c = ws.cell(row=1, column=1, value=f"📊 {deal} коммерческой недвижимости")
        c.font = Font(bold=True, size=14)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNS))
        ws.row_dimensions[1].height = 24
        for ci, name in enumerate(COLUMNS, start=1):
            cc = ws.cell(row=2, column=ci, value=name)
            cc.font = hdr_font
            cc.fill = hdr_fill
            cc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cc.border = cb
        ws.row_dimensions[2].height = 36
        deal_items = [it for it in items if it.get("_deal") == deal]
        deal_items.sort(
            key=lambda x: (type_idx.get(x.get("Тип"), 99), x.get("Дата публикации") or "")
        )
        row = 3
        prev_set = prev_hashes.get(deal, set())
        for it in deal_items:
            is_new = bool(prev_set) and str(it.get("Хэш", "")) not in prev_set
            if is_new:
                new_counts[deal] += 1
            for ci, name in enumerate(COLUMNS, start=1):
                val = it.get(name, "")
                cc = ws.cell(row=row, column=ci, value=val)
                cc.alignment = Alignment(vertical="top", wrap_text=True)
                cc.border = cb
                if name == "Дата публикации":
                    cc.number_format = "@"
                    cc.alignment = Alignment(horizontal="center", vertical="top")
                if name == "Тип":
                    cc.fill = PatternFill("solid", fgColor=TYPE_COLORS.get(val, "888888"))
                    cc.font = Font(bold=True, color="FFFFFF", size=10)
                    cc.alignment = Alignment(horizontal="center", vertical="center")
                elif name == "Ссылка" and val:
                    cc.hyperlink = val
                    cc.font = Font(color="0563C1", underline="single", size=10)
                elif name == "Телефон" and val:
                    cc.font = Font(bold=True, color="006100", size=11)
                elif val == "н/у":
                    cc.font = Font(color="999999", size=10)
                if is_new and name != "Тип":
                    cc.fill = new_fill
            row += 1
        # C3 = закрепляем шапку (строки 1-2) И столбцы «Сохранить»+«Тип» (A, B)
        ws.freeze_panes = "C3"
        last_col = get_column_letter(len(COLUMNS))
        ws.auto_filter.ref = f"A2:{last_col}{max(row - 1, 2)}"
        col_widths = {
            "Сохранить": 10,
            "Фото URL": 40,
            "Координаты": 18,
            "Тип": 11,
            "Телефон": 22,
            "Ссылка": 38,
            "Адрес": 28,
            "Район / Город": 18,
            "Площадь, м²": 12,
            "Цена общая": 22,
            "Цена за м²": 18,
            "НДС": 8,
            "Этаж / этажность": 13,
            "Год постройки": 10,
            "Класс здания": 11,
            "Состояние": 14,
            "Парковка": 10,
            "Отдельный вход": 14,
            "Мокрая зона": 12,
            "Контакт": 16,
            "Имя контакта": 16,
            "Дата публикации": 14,
            "Источник": 12,
            "Высота потолков, м": 14,
            "Грузовая рампа / ворота": 18,
            "Электр. мощность, кВт": 16,
            "Витринные окна / 1-я линия": 18,
            "Мин. срок аренды": 14,
            "Материал стен": 14,
            "Хэш": 14,
        }
        for ci, name in enumerate(COLUMNS, start=1):
            ws.column_dimensions[get_column_letter(ci)].width = col_widths.get(name, 14)
    summary = wb.create_sheet("Сводка", 0)
    summary.sheet_view.showGridLines = False
    summary["A1"] = "📅 Выгрузка коммерческой недвижимости"
    summary["A1"].font = Font(bold=True, size=16)
    summary["A2"] = f"Дата: {date.today().strftime('%d.%m.%Y')}"
    summary["A2"].font = Font(size=11, color="666666")
    phones_count = sum(1 for it in items if it.get("Телефон"))
    summary["A3"] = f"С телефонами: {phones_count}/{len(items)}"
    summary["A3"].font = Font(size=11, color="006100", bold=True)
    if prev_hashes:
        total_new = sum(new_counts.values())
        summary["A4"] = (
            f"🆕 Новых с прошлой выгрузки: {total_new} "
            f"(Продажа: {new_counts.get('Продажа', 0)}, "
            f"Аренда: {new_counts.get('Аренда', 0)}) — подсвечены жёлтым"
        )
        summary["A4"].font = Font(size=11, color="B86E00", bold=True)
    summary["A5"] = "Тип \\ Сделка"
    summary["B5"] = "Продажа"
    summary["C5"] = "Аренда"
    summary["D5"] = "Всего"
    for col in ["A5", "B5", "C5", "D5"]:
        summary[col].font = hdr_font
        summary[col].fill = hdr_fill
        summary[col].alignment = Alignment(horizontal="center")
        summary[col].border = cb
    for i, t in enumerate(TYPE_ORDER, start=6):
        sc = sum(1 for it in items if it["_deal"] == "Продажа" and it["Тип"] == t)
        ac = sum(1 for it in items if it["_deal"] == "Аренда" and it["Тип"] == t)
        summary[f"A{i}"] = t
        summary[f"A{i}"].fill = PatternFill("solid", fgColor=TYPE_COLORS[t])
        summary[f"A{i}"].font = Font(bold=True, color="FFFFFF")
        summary[f"B{i}"] = sc
        summary[f"C{i}"] = ac
        summary[f"D{i}"] = sc + ac
        for col in ["A", "B", "C", "D"]:
            summary[f"{col}{i}"].border = cb
            summary[f"{col}{i}"].alignment = Alignment(horizontal="center")
    tr = 6 + len(TYPE_ORDER)
    summary[f"A{tr}"] = "ИТОГО"
    summary[f"B{tr}"] = sum(1 for it in items if it["_deal"] == "Продажа")
    summary[f"C{tr}"] = sum(1 for it in items if it["_deal"] == "Аренда")
    summary[f"D{tr}"] = len(items)
    for col in ["A", "B", "C", "D"]:
        summary[f"{col}{tr}"].font = Font(bold=True, size=12)
        summary[f"{col}{tr}"].border = cb
        summary[f"{col}{tr}"].alignment = Alignment(horizontal="center")
    summary.column_dimensions["A"].width = 24
    for c in ["B", "C", "D"]:
        summary.column_dimensions[c].width = 14
    wb.save(output_path)
    msg = f"\n✅ Сохранено: {output_path}\n   Всего: {len(items)} | С телефонами: {phones_count}"
    if prev_hashes:
        msg += f" | 🆕 новых: {sum(new_counts.values())}"
    print(msg)


def parse_args() -> RunConfig:
    p = argparse.ArgumentParser(
        description="Парсер коммерческой недвижимости realt.by → Excel."
    )
    p.add_argument(
        "--limit",
        type=int,
        default=10_000,
        metavar="N",
        help="жёсткий потолок объявлений с категории (по умолчанию %(default)s — фактически без лимита)",
    )
    p.add_argument(
        "--max-pages",
        type=int,
        default=100,
        metavar="N",
        help="максимум страниц пагинации в категории (по умолчанию %(default)s)",
    )
    p.add_argument(
        "--full",
        action="store_true",
        help="полный перепрогон: игнорировать БД, обойти все страницы и перепарсить всё",
    )
    p.add_argument(
        "--no-details",
        action="store_true",
        help="не заходить в карточки (без телефонов и полей со страницы)",
    )
    p.add_argument(
        "--headed",
        action="store_true",
        help="браузер с окном (не headless)",
    )
    p.add_argument(
        "--out",
        type=Path,
        default=DEFAULT_OUT_FILE,
        help="путь к .xlsx (по умолчанию %(default)s)",
    )
    p.add_argument(
        "--debug",
        action="store_true",
        help=f"писать сбои коротких карточек / без телефона в {DEFAULT_DEBUG_FILE.name}",
    )
    p.add_argument(
        "--retries",
        type=int,
        default=3,
        metavar="N",
        help="повторов page.goto при сетевых сбоях (по умолчанию %(default)s)",
    )
    p.add_argument("-v", "--verbose", action="store_true", help="подробный вывод в консоль")
    ns = p.parse_args()
    return RunConfig(
        limit_per_category=max(1, ns.limit),
        fetch_details=not ns.no_details,
        headless=not ns.headed,
        out_file=ns.out.expanduser().resolve(),
        debug_file=(DEFAULT_DEBUG_FILE.resolve() if ns.debug else None),
        goto_retries=max(1, ns.retries),
        verbose=ns.verbose,
        max_pages=max(1, ns.max_pages),
        full_rescan=ns.full,
    )


async def collect_new(
    cfg: RunConfig,
    prev_urls: Set[str],
    last_run_date: Optional[date] = None,
    on_checkpoint: Optional[Callable[[list], None]] = None,
    checkpoint_every: int = 50,
) -> list[dict]:
    """Собирает realt.by и возвращает СПИСОК новых items (с деталями), НЕ записывая файл.

    Для оркестратора (collect_realty.py). Дедуп по prev_urls, инкрементальные
    стоп-условия как в scrape_category. Запись файла — забота вызывающего, но во
    время длинного обхода карточек каждые checkpoint_every вызывается on_checkpoint
    (список уже обработанных новых items) — для промежуточного сохранения.
    """
    if cfg.debug_file and not cfg.debug_file.exists():
        cfg.debug_file.write_text(
            f"# debug {date.today().isoformat()} (orchestrator)\n", encoding="utf-8"
        )
    all_items: list = []
    global_seen: Set[str] = set()
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=cfg.headless)
        context = await browser.new_context(
            user_agent=UA, viewport={"width": 1440, "height": 900}, locale="ru-RU"
        )
        page = await context.new_page()
        await apply_stealth(page, cfg)
        for i, cat in enumerate(CATEGORIES, 1):
            print(f"\n[realt {i}/{len(CATEGORIES)}]", end="")
            try:
                items = await scrape_category(
                    page, cat, cfg.limit_per_category, global_seen, cfg,
                    prev_urls=prev_urls if not cfg.full_rescan else None,
                    last_run_date=last_run_date if not cfg.full_rescan else None,
                )
                all_items.extend(items)
            except Exception as e:  # noqa: BLE001
                print(f"  ✖ {e}")
            await asyncio.sleep(random.uniform(1.5, 3))
        new_items = [it for it in all_items if normalize_url(it["Ссылка"]) not in prev_urls]
        print(f"\n  realt: листинг {len(all_items)} | новых {len(new_items)}")
        if cfg.fetch_details and new_items:
            detail_page = await context.new_page()
            await apply_stealth(detail_page, cfg)
            for i, item in enumerate(new_items, 1):
                try:
                    await fetch_details(detail_page, item, cfg)
                except Exception as e:  # noqa: BLE001
                    if cfg.verbose:
                        print(f"    деталь ✖ {e}")
                if i % 25 == 0:
                    print(f"    realt детали {i}/{len(new_items)}")
                if on_checkpoint and i % checkpoint_every == 0 and i < len(new_items):
                    try:
                        on_checkpoint(new_items[:i])
                        print(f"    💾 чекпойнт realt: {i}/{len(new_items)}")
                    except Exception as e:  # noqa: BLE001
                        print(f"    ⚠ чекпойнт не записан: {e}")
                await asyncio.sleep(random.uniform(2, 4))
            await detail_page.close()
        await browser.close()
    return new_items


async def run_scrape(cfg: RunConfig) -> None:
    if cfg.debug_file:
        cfg.debug_file.write_text(
            f"# debug {date.today().isoformat()} limit={cfg.limit_per_category}\n",
            encoding="utf-8",
        )
    # Загружаем БД из предыдущей выгрузки (если есть)
    prev_db, last_run_date = load_prev_db(cfg.out_file)
    prev_urls: Set[str] = set(prev_db.keys()) if not cfg.full_rescan else set()
    # Снапшот хэшей для подсветки новых — снимаем ДО любых чекпойнт-перезаписей файла.
    prev_hashes_snapshot: dict[str, set[str]] = {}
    for _url, _rec in prev_db.items():
        _deal, _h = _rec.get("_deal"), _rec.get("Хэш")
        if _deal and _h:
            prev_hashes_snapshot.setdefault(_deal, set()).add(str(_h))
    mode = "ПОЛНЫЙ ПЕРЕПРОГОН" if cfg.full_rescan else (
        f"инкрементальный (в БД {len(prev_db)} URL"
        + (f", last_run={last_run_date:%d.%m.%Y}" if last_run_date else "")
        + ")"
        if prev_db else "первый прогон (БД пуста)"
    )
    print(f"🚀 realty_parser_v8. Дата: {date.today():%d.%m.%Y}")
    print(f"   режим: {mode}")
    print(
        f"   детали={cfg.fetch_details}, max_pages={cfg.max_pages}, "
        f"limit/cat={cfg.limit_per_category}, out={cfg.out_file}"
    )
    if cfg.debug_file:
        print(f"   debug → {cfg.debug_file}")
    all_items: list = []
    global_seen: Set[str] = set()
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=cfg.headless)
        context = await browser.new_context(
            user_agent=UA,
            viewport={"width": 1440, "height": 900},
            locale="ru-RU",
        )
        page = await context.new_page()
        await apply_stealth(page, cfg)
        for i, cat in enumerate(CATEGORIES, 1):
            print(f"\n[{i}/{len(CATEGORIES)}]", end="")
            try:
                items = await scrape_category(
                    page, cat, cfg.limit_per_category, global_seen, cfg,
                    prev_urls=prev_urls if not cfg.full_rescan else None,
                    last_run_date=last_run_date if not cfg.full_rescan else None,
                )
                all_items.extend(items)
            except Exception as e:
                print(f"  ✖ {e}")
            await asyncio.sleep(random.uniform(1.5, 3))
        # Разделяем на новые и известные
        new_items = [it for it in all_items if normalize_url(it["Ссылка"]) not in prev_urls]
        print(
            f"\n\n📋 Из листинга собрано: {len(all_items)} | "
            f"новых: {len(new_items)} | известных: {len(all_items) - len(new_items)}"
        )
        # Финальный список = старая БД (новые URL с ней не пересекаются) + свежие.
        def build_final(done_new: list) -> list:
            base = [] if cfg.full_rescan else list(prev_db.values())
            return base + list(done_new)

        done_new: list = []
        if cfg.fetch_details and new_items:
            print(f"\n🔍 Загружаю детали для {len(new_items)} новых...")
            detail_page = await context.new_page()
            await apply_stealth(detail_page, cfg)
            for i, item in enumerate(new_items, 1):
                addr = (item.get("Адрес") or item.get("Ссылка") or "")[:50]
                print(f"  [{i}/{len(new_items)}] {addr}...", end=" ")
                try:
                    await fetch_details(detail_page, item, cfg)
                    if item.get("Телефон"):
                        print(f"📞 {item['Телефон'][:20]}")
                    else:
                        print("без тел.")
                except Exception as e:
                    print(f"✖ {e}")
                done_new.append(item)
                # Промежуточное сохранение: защита от потери данных при сбое/Ctrl-C.
                if i % CHECKPOINT_EVERY == 0 and i < len(new_items):
                    try:
                        write_excel(
                            build_final(done_new), cfg.out_file,
                            prev_hashes=prev_hashes_snapshot,
                        )
                        print(f"    💾 чекпойнт: {len(done_new)} новых сохранено")
                    except Exception as e:
                        print(f"    ⚠ чекпойнт не записан: {e}")
                await asyncio.sleep(random.uniform(2, 4))
            await detail_page.close()
        elif not new_items:
            print("\n✨ Новых объявлений нет, детали не нужны.")
        else:
            done_new = list(new_items)
        await browser.close()
    final_items = build_final(done_new)
    print(
        f"\n📦 Итог: {len(final_items)} объектов "
        f"(новых: {len(done_new)}, из БД: {len(final_items) - len(done_new)})"
    )
    write_excel(final_items, cfg.out_file, prev_hashes=prev_hashes_snapshot)


def main() -> None:
    cfg = parse_args()
    asyncio.run(run_scrape(cfg))


if __name__ == "__main__":
    main()
