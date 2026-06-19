#!/usr/bin/env python3
"""Сбор телефонов компаний с belretail.by по категориям → belretail_phones.xlsx.

⚠️ ЗАПУСКАТЬ С БЕЛОРУССКИМ IP (Psiphon ВЫКЛ) — belretail не отвечает на иностранный IP.

Каждая страница belretail.by/retailcategories/{slug} — список компаний (блоки
div.retail-item: имя в span.name, телефон в тексте «Телефон: …»). Скрипт обходит все
17 категорий, парсит и пишет xlsx, сгруппированный по категориям и компаниям.

  ./bin/python belretail_phones.py
"""
import re
import time
import urllib.request
from pathlib import Path

from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

HERE = Path(__file__).resolve().parent
OUT = HERE / "belretail_phones.xlsx"
UA = ("Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
      "(KHTML, like Gecko) Chrome/120.0 Safari/537.36")
BASE = "https://belretail.by/retailcategories/"

# слаг → читаемое имя (17 категорий из индекса /retailcategories)
CATEGORIES = {
    "produktyi": "Продукты", "alco": "Алкоголь", "apteki": "Аптеки", "avto": "Авто",
    "detskie-tovaryi": "Детские товары", "diy-stroymarketyi": "DIY / Строймаркеты",
    "elektronika": "Электроника", "kafe-restoranyi": "Кафе и рестораны", "knigi": "Книги",
    "kosmetika-parfyumeriya": "Косметика и парфюмерия", "mebel": "Мебель",
    "odejda-obuv-aksessuaryi": "Одежда, обувь, аксессуары", "podarki-tsvetyi": "Подарки и цветы",
    "products-for-animals": "Товары для животных", "sotovaya-svyaz": "Сотовая связь",
    "sportivnyie-tovaryi": "Спортивные товары", "yuvelirnyie-izdeliya": "Ювелирные изделия",
}


def fetch(url):
    req = urllib.request.Request(url, headers={"User-Agent": UA})
    with urllib.request.urlopen(req, timeout=30) as r:
        return r.read().decode("utf-8", "replace")


def phones_from(text):
    """Все номера после «Телефон:»; +375.. и 8(0xx).. → канон +375XXXXXXXXX; прочее как есть."""
    m = re.search(r"Телефон[:\s]*(.+)", text)
    seg = m.group(1) if m else ""
    out = []
    for r in re.findall(r"(?:\+?375|8)[\d\s\-()]{7,}", seg):
        d = re.sub(r"\D", "", r)
        if d.startswith("375") and len(d) >= 12:
            out.append("+" + d[:12])
        elif d.startswith("80") and len(d) >= 11:
            out.append("+375" + d[2:11])
        else:
            out.append(r.strip())
    return list(dict.fromkeys(out))   # дедуп, порядок сохранён


def parse_category(html):
    """[(компания, телефоны-строкой, ссылка)] из HTML страницы категории."""
    soup = BeautifulSoup(html, "lxml")
    rows = []
    for it in soup.select("div.retail-item"):
        a = it.select_one("span.name a")
        if not a:
            continue
        rows.append((a.get_text(strip=True),
                     ", ".join(phones_from(it.get_text(" ", strip=True))),
                     a.get("href", "")))
    return rows


def main():
    wb = Workbook()
    ws = wb.active
    ws.title = "Телефоны belretail"
    ws.append(["Категория", "Компания", "Телефон", "Ссылка"])
    for c in ws[1]:
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="DDDDDD")

    total = with_ph = 0
    for slug, name in CATEGORIES.items():
        try:
            rows = parse_category(fetch(BASE + slug))
        except Exception as e:  # noqa: BLE001 — категория может не открыться, продолжаем
            print(f"  ⚠ {name}: {e}")
            continue
        for company, phone, link in rows:
            ws.append([name, company, phone, link])
            total += 1
            with_ph += bool(phone)
        print(f"  {name}: {len(rows)} компаний")
        time.sleep(1.5)   # вежливая пауза между категориями

    for col, w in zip("ABCD", (24, 34, 28, 52)):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"
    wb.save(OUT)
    print(f"\n✅ {total} компаний ({with_ph} с телефоном) → {OUT}")


if __name__ == "__main__":
    main()
